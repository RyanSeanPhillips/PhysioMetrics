"""
File Index Service — scan directories, classify files, extract notes text, search.

Pure Python, no Qt dependencies.
"""

import os
from pathlib import Path
from datetime import datetime, timezone
from typing import List, Optional, Dict, Any

from core.adapters.file_index_store import FileIndexStore


# Extension → file_class mapping
DATA_EXTENSIONS = {".abf", ".smrx", ".edf", ".mat", ".npz"}
VIDEO_EXTENSIONS = {".mp4", ".avi", ".mov", ".mkv"}
NOTES_KEYWORDS = {"note", "notes"}
REFERENCE_KEYWORDS = {"mice", "mouse", "subject", "animal", "master"}
SPREADSHEET_EXTENSIONS = {".xlsx", ".xls", ".csv"}

# Directories to skip during scanning
SKIP_DIRS = {
    ".git", "__pycache__", "node_modules", ".venv", "venv",
    "$RECYCLE.BIN", "System Volume Information",
}

# Max directory depth to prevent runaway scans
DEFAULT_MAX_DEPTH = 8


class FileIndexService:
    """Service for building and querying the lab file index."""

    def __init__(self, store: Optional[FileIndexStore] = None):
        self.store = store or FileIndexStore()

    def close(self):
        self.store.close()

    # --- Classification ---

    @staticmethod
    def classify_file(file_name: str, extension: str) -> str:
        """Classify a file by extension and name patterns."""
        ext = extension.lower()
        name_lower = file_name.lower()

        if ext in DATA_EXTENSIONS:
            return "data"

        if ext == ".csv":
            # CSVs can be data (photometry) or notes — check name
            if any(kw in name_lower for kw in NOTES_KEYWORDS):
                return "notes"
            if any(kw in name_lower for kw in REFERENCE_KEYWORDS):
                return "reference"
            return "data"

        if ext in SPREADSHEET_EXTENSIONS:
            if any(kw in name_lower for kw in NOTES_KEYWORDS):
                return "notes"
            if any(kw in name_lower for kw in REFERENCE_KEYWORDS):
                return "reference"
            return "other"

        if ext in VIDEO_EXTENSIONS:
            return "video"

        return "other"

    # --- Scanning ---

    def scan_directory(self, root_path: str, recursive: bool = True,
                       max_depth: int = DEFAULT_MAX_DEPTH,
                       progress_callback=None) -> Dict[str, Any]:
        """Walk directory tree, classify and index all files. Returns summary."""
        root = Path(root_path)
        if not root.exists():
            return {"error": f"Path does not exist: {root_path}"}

        counts = {"added": 0, "updated": 0, "skipped": 0, "deleted": 0}
        class_counts: Dict[str, int] = {}
        seen_paths: set = set()
        root_depth = len(root.parts)

        for dirpath, dirnames, filenames in os.walk(str(root)):
            # Skip hidden/system dirs
            dirnames[:] = [d for d in dirnames if d not in SKIP_DIRS and not d.startswith(".")]

            # Depth check
            current_depth = len(Path(dirpath).parts) - root_depth
            if not recursive and current_depth > 0:
                dirnames.clear()
                continue
            if current_depth > max_depth:
                dirnames.clear()
                continue

            for fname in filenames:
                fpath = Path(dirpath) / fname
                try:
                    stat = fpath.stat()
                except (OSError, PermissionError):
                    counts["skipped"] += 1
                    continue

                ext = fpath.suffix.lower()
                file_path_str = str(fpath)
                seen_paths.add(file_path_str)

                # Check if already indexed with same mtime
                existing = self.store.get_file(file_path_str)
                mtime_iso = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat()

                if existing and existing["mtime"] == mtime_iso:
                    counts["skipped"] += 1
                    continue

                file_class = self.classify_file(fname, ext)
                self.store.upsert_file(
                    file_path=file_path_str,
                    file_name=fname,
                    parent_dir=fpath.parent.name,
                    extension=ext,
                    file_size=stat.st_size,
                    mtime=mtime_iso,
                    file_class=file_class,
                )

                if existing:
                    counts["updated"] += 1
                else:
                    counts["added"] += 1
                class_counts[file_class] = class_counts.get(file_class, 0) + 1

                if progress_callback and (counts["added"] + counts["updated"]) % 100 == 0:
                    progress_callback(counts["added"] + counts["updated"])

        # Clean up deleted files
        counts["deleted"] = self.store.delete_missing_files(seen_paths)

        # Save scan metadata
        now = datetime.now(timezone.utc).isoformat()
        self.store.set_meta("last_scan", now)
        # Append to scan roots
        existing_roots = self.store.get_meta("scan_roots") or ""
        root_str = str(root)
        if root_str not in existing_roots:
            roots = [r for r in existing_roots.split("|") if r] + [root_str]
            self.store.set_meta("scan_roots", "|".join(roots))

        return {
            "root": str(root),
            "added": counts["added"],
            "updated": counts["updated"],
            "skipped": counts["skipped"],
            "deleted": counts["deleted"],
            "by_class": class_counts,
        }

    # --- Text Extraction ---

    def extract_notes(self, file_ids: Optional[List[int]] = None,
                      force: bool = False,
                      progress_callback=None) -> Dict[str, Any]:
        """Extract text from notes/reference files into local cache.

        If file_ids is None, extracts all notes/reference files.
        Incremental by default: skips files whose cells are already cached
        unless force=True or file mtime has changed.
        """
        if file_ids:
            files = [self.store.get_file_by_id(fid) for fid in file_ids]
            files = [f for f in files if f is not None]
        else:
            files = self.store.get_files(file_class="notes") + self.store.get_files(file_class="reference")

        results = {"extracted": 0, "skipped": 0, "errors": [], "total_cells": 0}

        for i, f in enumerate(files):
            file_id = f["file_id"]
            file_path = f["file_path"]

            # Incremental: skip if already extracted and mtime unchanged
            if not force and self.store.has_cells(file_id):
                results["skipped"] += 1
                continue

            try:
                cells = self._read_file_cells(file_path, f["extension"])
                if cells:
                    self.store.upsert_cells(file_id, cells)
                    results["extracted"] += 1
                    results["total_cells"] += len(cells)
                else:
                    results["skipped"] += 1
            except Exception as e:
                results["errors"].append({"file": file_path, "error": str(e)})

            if progress_callback:
                progress_callback(i + 1, len(files), f["file_name"])

        now = datetime.now(timezone.utc).isoformat()
        self.store.set_meta("last_extract", now)
        return results

    def _read_file_cells(self, file_path: str, extension: str) -> List[Dict[str, Any]]:
        """Read all cell text from a spreadsheet file."""
        ext = extension.lower()
        if ext in (".xlsx", ".xls"):
            return self._read_excel_cells(file_path)
        elif ext == ".csv":
            return self._read_csv_cells(file_path)
        return []

    @staticmethod
    def _read_excel_cells(file_path: str) -> List[Dict[str, Any]]:
        """Read all cells from an Excel file using openpyxl."""
        import openpyxl
        cells = []
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception:
            # Try without read_only for older formats
            wb = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
                for cell in row:
                    if cell.value is not None:
                        val = str(cell.value).strip()
                        if val:
                            cells.append({
                                "sheet_name": sheet_name,
                                "row_num": row_idx,
                                "col_num": cell.column,
                                "value": val,
                            })
        wb.close()
        return cells

    @staticmethod
    def _read_csv_cells(file_path: str) -> List[Dict[str, Any]]:
        """Read all cells from a CSV file."""
        import csv
        cells = []
        # Try common encodings
        for encoding in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
            try:
                with open(file_path, "r", encoding=encoding, newline="") as f:
                    reader = csv.reader(f)
                    for row_idx, row in enumerate(reader, start=1):
                        for col_idx, val in enumerate(row, start=1):
                            val = val.strip()
                            if val:
                                cells.append({
                                    "sheet_name": "",
                                    "row_num": row_idx,
                                    "col_num": col_idx,
                                    "value": val,
                                })
                return cells
            except UnicodeDecodeError:
                continue
        return cells

    # --- Search ---

    def search(self, query: str, scope: str = "all",
               limit: int = 100) -> Dict[str, Any]:
        """Unified search across files and notes content."""
        results: Dict[str, Any] = {}

        if scope in ("all", "files"):
            results["files"] = self.store.search_files(query, limit=limit)

        if scope in ("all", "notes"):
            results["notes"] = self.store.search_notes(query, limit=limit)

        return results

    def find_file_in_notes(self, file_name: str,
                           limit: int = 50) -> List[Dict[str, Any]]:
        """Find which notes files mention a specific experiment file name.

        Key use case: given "25121003.abf", find the notes file + sheet + row.
        Strips extension for broader matching.
        """
        # Try without extension first
        stem = Path(file_name).stem
        matches = self.store.find_value_in_notes(stem, limit=limit)

        # Group by source file for cleaner output
        grouped: Dict[str, Dict[str, Any]] = {}
        for m in matches:
            key = f"{m['file_path']}|{m['sheet_name']}"
            if key not in grouped:
                grouped[key] = {
                    "source_file": m["file_path"],
                    "source_name": m["file_name"],
                    "sheet": m["sheet_name"],
                    "matches": [],
                }
            grouped[key]["matches"].append({
                "row": m["row_num"],
                "col": m["col_num"],
                "value": m["value"],
            })

        return list(grouped.values())

    def get_notes_summary(self, file_id: int) -> Dict[str, Any]:
        """Get summary of a notes file: sheets, row counts, detected headers."""
        f = self.store.get_file_by_id(file_id)
        if not f:
            return {"error": "File not found"}

        sheets = self.store.get_sheets(file_id)
        summary = {
            "file_path": f["file_path"],
            "file_name": f["file_name"],
            "sheets": [],
        }

        for sheet in sheets:
            cells = self.store.get_cells(file_id, sheet_name=sheet)
            if not cells:
                continue

            max_row = max(c["row_num"] for c in cells)
            max_col = max(c["col_num"] for c in cells)

            # Detect header row (first row with 3+ non-empty cells)
            rows_data: Dict[int, List[str]] = {}
            for c in cells:
                rows_data.setdefault(c["row_num"], []).append(c["value"])

            header_row = None
            headers = []
            for rn in sorted(rows_data.keys()):
                if len(rows_data[rn]) >= 3:
                    header_row = rn
                    # Get ordered header values
                    header_cells = [c for c in cells if c["row_num"] == rn]
                    header_cells.sort(key=lambda c: c["col_num"])
                    headers = [c["value"] for c in header_cells]
                    break

            summary["sheets"].append({
                "name": sheet,
                "rows": max_row,
                "cols": max_col,
                "cell_count": len(cells),
                "header_row": header_row,
                "headers": headers,
            })

        return summary

    def get_stats(self) -> Dict[str, Any]:
        return self.store.get_stats()

    def reclassify_file(self, file_path: str, new_class: str) -> bool:
        """Manually override a file's classification."""
        f = self.store.get_file(file_path)
        if not f:
            return False
        self.store.upsert_file(
            file_path=f["file_path"],
            file_name=f["file_name"],
            parent_dir=f["parent_dir"],
            extension=f["extension"],
            file_size=f["file_size"],
            mtime=f["mtime"],
            file_class=new_class,
        )
        return True
