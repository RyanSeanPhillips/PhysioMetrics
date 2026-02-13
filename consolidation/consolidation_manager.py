"""
ConsolidationManager - Handles multi-file data consolidation and Excel export.

Extracted from main.py for better maintainability and easier customization
for different experiment types.
"""

import re
import numpy as np
import pandas as pd
from pathlib import Path
from PyQt6.QtWidgets import QMessageBox, QFileDialog, QProgressDialog, QApplication
from PyQt6.QtCore import Qt


class ConsolidationManager:
    """Manages all multi-file consolidation operations for the main window."""

    def __init__(self, main_window):
        """
        Initialize the ConsolidationManager.

        Args:
            main_window: Reference to MainWindow instance
        """
        self.window = main_window

    def _propose_consolidated_filename(self, files: list) -> tuple[str, list[str]]:
        """Generate a descriptive filename based on the files being consolidated.

        Returns:
            tuple: (proposed_filename, list of warnings)
        """
        warnings = []

        if not files:
            return "consolidated_data.xlsx", warnings

        # Parse filenames to extract common metadata
        # Expected format: Strain_Virus_Location_Sex_Animal_Stim_Power_ABF_Channel_*.csv
        # Also support older format without Location
        parsed_files = []
        for _, filepath in files:
            stem = filepath.stem  # filename without extension
            parts = stem.split('_')

            # Try to intelligently parse - check if we have location field
            # Location field is typically anatomical abbreviations like preBotC, RTN, etc.
            # Sex is typically M, F, or Unknown (short)
            # We'll use heuristics: if part[2] is 1-2 chars, it's probably sex (old format)

            if len(parts) >= 7 and len(parts[3]) <= 2:
                # New format with location: Strain_Virus_Location_Sex_Animal_Stim_Power_...
                parsed_files.append({
                    'strain': parts[0] if len(parts) > 0 else '',
                    'virus': parts[1] if len(parts) > 1 else '',
                    'location': parts[2] if len(parts) > 2 else '',
                    'sex': parts[3] if len(parts) > 3 else '',
                    'animal': parts[4] if len(parts) > 4 else '',
                    'stim': parts[5] if len(parts) > 5 else '',
                    'power': parts[6] if len(parts) > 6 else '',
                })
            elif len(parts) >= 6 and len(parts[2]) <= 2:
                # Old format without location: Strain_Virus_Sex_Animal_Stim_Power_...
                parsed_files.append({
                    'strain': parts[0] if len(parts) > 0 else '',
                    'virus': parts[1] if len(parts) > 1 else '',
                    'location': '',
                    'sex': parts[2] if len(parts) > 2 else '',
                    'animal': parts[3] if len(parts) > 3 else '',
                    'stim': parts[4] if len(parts) > 4 else '',
                    'power': parts[5] if len(parts) > 5 else '',
                })

        if not parsed_files:
            return "consolidated_data.xlsx", warnings

        # Find common values across all files and check for variations
        common = {}
        for key in ['strain', 'virus', 'location', 'sex', 'stim', 'power']:
            values = set(f[key] for f in parsed_files if f.get(key))
            if len(values) == 1:
                common[key] = values.pop()
            elif len(values) > 1:
                # Warn about different stimulation parameters
                if key == 'stim':
                    warnings.append(f"Multiple stimulation types detected: {', '.join(sorted(values))}")
                elif key == 'power':
                    warnings.append(f"Multiple laser powers detected: {', '.join(sorted(values))}")

        # Build descriptive filename from common fields
        parts = []
        if common.get('strain'):
            parts.append(common['strain'])
        if common.get('virus'):
            parts.append(common['virus'])
        if common.get('location'):
            parts.append(common['location'])
        if common.get('sex'):
            parts.append(common['sex'])

        # If multiple animals, indicate that
        animals = set(f['animal'] for f in parsed_files if f.get('animal'))
        if len(animals) == 1:
            parts.append(animals.pop())
        elif len(animals) > 1:
            parts.append(f"N{len(animals)}")  # Capital N

        if common.get('stim'):
            parts.append(common['stim'])
        if common.get('power'):
            parts.append(common['power'])

        # Add "consolidated" suffix
        parts.append("consolidated")

        if parts:
            return "_".join(parts) + ".xlsx", warnings
        else:
            return "consolidated_data.xlsx", warnings

    def consolidate_csv_files(self, means_files: list, breaths_files: list, events_files: list):
        """
        Consolidate CSV files directly (used by Project Builder Consolidation tab).

        Args:
            means_files: List of (root_name, Path) tuples for timeseries/means CSVs
            breaths_files: List of (root_name, Path) tuples for breaths CSVs
            events_files: List of (root_name, Path) tuples for events CSVs
        """
        if not means_files and not breaths_files and not events_files:
            QMessageBox.warning(self.window, "Consolidate", "No CSV files to consolidate.")
            return

        # Choose save location with intelligent default name
        files_for_naming = means_files or breaths_files
        proposed_filename, warnings = self._propose_consolidated_filename(files_for_naming)

        # Show warnings if any
        if warnings:
            warning_msg = "Warning about files being consolidated:\n\n" + "\n".join(f"• {w}" for w in warnings)
            warning_msg += "\n\nDo you want to continue?"
            reply = QMessageBox.question(
                self.window, "Consolidation Warning",
                warning_msg,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return

        # Determine default save location
        if means_files:
            default_name = str(means_files[0][1].parent / proposed_filename)
        elif breaths_files:
            default_name = str(breaths_files[0][1].parent / proposed_filename)
        else:
            default_name = proposed_filename

        save_path, _ = QFileDialog.getSaveFileName(
            self.window, "Save consolidated data as...",
            default_name,
            "Excel Files (*.xlsx)"
        )

        if not save_path:
            return

        # Create progress dialog
        progress = QProgressDialog("Consolidating data...", "Cancel", 0, 100, self.window)
        progress.setWindowTitle("PhysioMetrics")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)

        try:
            consolidated_data = {}

            if means_files:
                progress.setLabelText(f"Processing time series data ({len(means_files)} files)...")
                progress.setValue(10)
                QApplication.processEvents()
                if progress.wasCanceled():
                    return
                consolidated_data.update(self._consolidate_means_files(means_files))
                progress.setValue(40)

            if breaths_files:
                if progress.wasCanceled():
                    return
                progress.setLabelText(f"Processing breath histograms ({len(breaths_files)} files)...")
                progress.setValue(50)
                QApplication.processEvents()
                histogram_data = self._consolidate_breaths_histograms(breaths_files)
                consolidated_data.update(histogram_data)
                progress.setValue(70)

                # Extract sigh data
                progress.setLabelText("Extracting sigh data...")
                sighs_df = self._consolidate_breaths_sighs(breaths_files)
                consolidated_data['sighs'] = {
                    'time_series': sighs_df,
                    'raw_summary': {},
                    'norm_summary': {},
                    'windows': []
                }
                progress.setValue(80)

            if events_files:
                if progress.wasCanceled():
                    return
                progress.setLabelText(f"Processing events data ({len(events_files)} files)...")
                progress.setValue(82)
                QApplication.processEvents()
                events_df = self._consolidate_events(events_files)
                consolidated_data['events'] = {
                    'time_series': events_df,
                    'raw_summary': {},
                    'norm_summary': {},
                    'eupnea_summary': {},
                    'windows': []
                }

                # Process stimulus events
                progress.setLabelText("Processing stimulus data...")
                progress.setValue(83)
                QApplication.processEvents()
                stimulus_df, stim_warnings = self._consolidate_stimulus(events_files)
                consolidated_data['stimulus'] = {
                    'time_series': stimulus_df,
                    'raw_summary': {},
                    'norm_summary': {},
                    'windows': []
                }

            progress.setLabelText("Writing Excel file...")
            progress.setValue(90)
            QApplication.processEvents()

            # Write to Excel
            self._write_consolidated_excel(save_path, consolidated_data)

            progress.setValue(100)
            progress.close()

            QMessageBox.information(
                self.window, "Success",
                f"Consolidated data saved to:\n{save_path}"
            )

        except Exception as e:
            progress.close()
            QMessageBox.critical(
                self.window, "Error",
                f"Failed to consolidate data:\n{str(e)}"
            )
            import traceback
            traceback.print_exc()

    def consolidate_files(self, files: list):
        """
        Consolidate NPZ files from the Project Builder Consolidation tab.

        Args:
            files: List of (display_name, npz_path) tuples
        """
        if not files:
            QMessageBox.warning(self.window, "Consolidate", "No files selected to consolidate.")
            return

        # Find associated CSV files for each NPZ
        means_files = []
        breaths_files = []
        events_files = []

        for display_name, npz_path in files:
            npz_path = Path(npz_path)
            if not npz_path.exists():
                continue

            # Look for associated CSV files in the same directory
            base_dir = npz_path.parent
            base_name = npz_path.stem

            # Try common naming patterns for CSV files
            means_candidates = [
                base_dir / f"{base_name}_means.csv",
                base_dir / f"{base_name}.csv",
            ]
            breaths_candidates = [
                base_dir / f"{base_name}_breaths.csv",
            ]
            events_candidates = [
                base_dir / f"{base_name}_events.csv",
            ]

            for means_path in means_candidates:
                if means_path.exists():
                    means_files.append((display_name, means_path))
                    break

            for breaths_path in breaths_candidates:
                if breaths_path.exists():
                    breaths_files.append((display_name, breaths_path))
                    break

            for events_path in events_candidates:
                if events_path.exists():
                    events_files.append((display_name, events_path))
                    break

        if not means_files and not breaths_files and not events_files:
            QMessageBox.warning(
                self.window, "Consolidate",
                "No CSV files found for selected NPZ files.\n\n"
                "Please ensure the analysis data has been exported to CSV format first."
            )
            return

        # Now use the same consolidation logic
        files_for_naming = means_files or breaths_files
        proposed_filename, warnings = self._propose_consolidated_filename(files_for_naming)

        if warnings:
            warning_msg = "Warning about files being consolidated:\n\n" + "\n".join(f"• {w}" for w in warnings)
            warning_msg += "\n\nDo you want to continue?"
            reply = QMessageBox.question(
                self.window, "Consolidation Warning",
                warning_msg,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return

        if means_files:
            default_name = str(means_files[0][1].parent / proposed_filename)
        elif breaths_files:
            default_name = str(breaths_files[0][1].parent / proposed_filename)
        else:
            default_name = proposed_filename

        save_path, _ = QFileDialog.getSaveFileName(
            self.window, "Save consolidated data as...",
            default_name,
            "Excel Files (*.xlsx)"
        )

        if not save_path:
            return

        # Create progress dialog
        n_total_files = len(means_files) + len(breaths_files)
        progress = QProgressDialog("Consolidating data...", "Cancel", 0, 100, self.window)
        progress.setWindowTitle("PhysioMetrics")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)

        try:
            consolidated_data = {}

            if means_files:
                progress.setLabelText(f"Processing time series data ({len(means_files)} files)...")
                progress.setValue(10)
                QApplication.processEvents()
                if progress.wasCanceled():
                    return
                consolidated_data.update(self._consolidate_means_files(means_files))
                progress.setValue(40)

            if breaths_files:
                if progress.wasCanceled():
                    return
                progress.setLabelText(f"Processing breath histograms ({len(breaths_files)} files)...")
                progress.setValue(50)
                QApplication.processEvents()
                histogram_data = self._consolidate_breaths_histograms(breaths_files)
                consolidated_data.update(histogram_data)
                progress.setValue(70)

                # Extract sigh data
                progress.setLabelText("Extracting sigh data...")
                sighs_df = self._consolidate_breaths_sighs(breaths_files)
                consolidated_data['sighs'] = {
                    'time_series': sighs_df,
                    'raw_summary': {},
                    'norm_summary': {},
                    'windows': []
                }
                progress.setValue(80)

            if events_files:
                if progress.wasCanceled():
                    return
                progress.setLabelText(f"Processing events data ({len(events_files)} files)...")
                progress.setValue(82)
                QApplication.processEvents()
                events_df = self._consolidate_events(events_files)
                consolidated_data['events'] = {
                    'time_series': events_df,
                    'raw_summary': {},
                    'norm_summary': {},
                    'eupnea_summary': {},
                    'windows': []
                }

                # Process stimulus events
                progress.setLabelText(f"Processing stimulus data...")
                progress.setValue(83)
                QApplication.processEvents()
                stimulus_df, stim_warnings = self._consolidate_stimulus(events_files)
                consolidated_data['stimulus'] = {
                    'time_series': stimulus_df,
                    'raw_summary': {},
                    'norm_summary': {},
                    'windows': []
                }

            progress.setLabelText("Writing Excel file...")
            progress.setValue(90)
            QApplication.processEvents()

            # Write to Excel
            self._write_consolidated_excel(save_path, consolidated_data)

            progress.setValue(100)
            progress.close()

            QMessageBox.information(
                self.window, "Success",
                f"Consolidated data saved to:\n{save_path}"
            )

        except Exception as e:
            progress.close()
            QMessageBox.critical(
                self.window, "Error",
                f"Failed to consolidate data:\n{str(e)}"
            )
            import traceback
            traceback.print_exc()

    def on_consolidate_save_data_clicked(self):
        """Consolidate data from selected files into a single Excel file."""
        from PyQt6.QtCore import Qt
        from PyQt6.QtWidgets import QProgressDialog
        import pandas as pd
        from pathlib import Path

        # Get all selected files from right list
        items = []
        for i in range(self.window.FilestoConsolidateList.count()):
            item = self.window.FilestoConsolidateList.item(i)
            if item:
                items.append(item)

        if not items:
            QMessageBox.warning(self.window, "Consolidate", "No files selected to consolidate.")
            return

        # Separate by file type
        means_files = []
        breaths_files = []
        events_files = []

        for item in items:
            meta = item.data(Qt.ItemDataRole.UserRole) or {}
            if meta.get("means"):
                means_files.append((meta["root"], Path(meta["means"])))
            if meta.get("breaths"):
                breaths_files.append((meta["root"], Path(meta["breaths"])))
            if meta.get("events"):
                events_files.append((meta["root"], Path(meta["events"])))

        if not means_files and not breaths_files and not events_files:
            QMessageBox.warning(self.window, "Consolidate", "No CSV files selected.")
            return

        # Choose save location first with intelligent default name
        files_for_naming = means_files or breaths_files
        proposed_filename, warnings = self._propose_consolidated_filename(files_for_naming)

        # Show warnings if any
        if warnings:
            warning_msg = "Warning about files being consolidated:\n\n" + "\n".join(f"• {w}" for w in warnings)
            warning_msg += "\n\nDo you want to continue?"
            reply = QMessageBox.question(
                self.window, "Consolidation Warning",
                warning_msg,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return

        if means_files:
            default_name = str(means_files[0][1].parent / proposed_filename)
        elif breaths_files:
            default_name = str(breaths_files[0][1].parent / proposed_filename)
        else:
            default_name = proposed_filename

        save_path, _ = QFileDialog.getSaveFileName(
            self.window, "Save consolidated data as...",
            default_name,
            "Excel Files (*.xlsx)"
        )

        if not save_path:
            return

        # Create progress dialog
        n_total_files = len(means_files) + len(breaths_files)
        progress = QProgressDialog("Consolidating data...", "Cancel", 0, 100, self.window)
        progress.setWindowTitle("PhysioMetrics")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)

        # Process files
        try:
            consolidated_data = {}

            if means_files:
                progress.setLabelText(f"Processing time series data ({len(means_files)} files)...")
                progress.setValue(10)
                QApplication.processEvents()
                if progress.wasCanceled():
                    return

                consolidated_data.update(self._consolidate_means_files(means_files))
                progress.setValue(40)
                QApplication.processEvents()

            if breaths_files:
                if progress.wasCanceled():
                    return

                progress.setLabelText(f"Processing breath histograms ({len(breaths_files)} files)...")
                progress.setValue(50)
                QApplication.processEvents()

                histogram_data = self._consolidate_breaths_histograms(breaths_files)
                consolidated_data.update(histogram_data)

                progress.setValue(70)
                QApplication.processEvents()
                if progress.wasCanceled():
                    return

                # Extract sigh data
                progress.setLabelText("Extracting sigh data...")
                sighs_df = self._consolidate_breaths_sighs(breaths_files)
                consolidated_data['sighs'] = {
                    'time_series': sighs_df,
                    'raw_summary': {},
                    'norm_summary': {},
                    'windows': []
                }
                progress.setValue(80)
                QApplication.processEvents()

            if events_files:
                if progress.wasCanceled():
                    return

                progress.setLabelText(f"Processing events data ({len(events_files)} files)...")
                progress.setValue(82)
                QApplication.processEvents()

                events_df = self._consolidate_events(events_files)
                print(f"Events DataFrame shape: {events_df.shape}")
                print(f"Events columns: {events_df.columns.tolist()}")
                if len(events_df) > 0:
                    print(f"First few event types: {events_df['event_type'].unique()[:5]}")
                consolidated_data['events'] = {
                    'time_series': events_df,
                    'raw_summary': {},
                    'norm_summary': {},
                    'eupnea_summary': {},
                    'windows': []
                }

                # Process stimulus events separately
                progress.setLabelText(f"Processing stimulus data ({len(events_files)} files)...")
                progress.setValue(83)
                QApplication.processEvents()

                stimulus_df, stim_warnings = self._consolidate_stimulus(events_files)
                print(f"Stimulus DataFrame shape: {stimulus_df.shape}")
                if len(stimulus_df) > 0:
                    print(f"Stimulus columns: {stimulus_df.columns.tolist()}")
                    print(f"Stimulus rows: {len(stimulus_df)}")
                consolidated_data['stimulus'] = {
                    'time_series': stimulus_df,
                    'raw_summary': {},
                    'norm_summary': {},
                    'eupnea_summary': {},
                    'windows': []
                }

                # Add stimulus warnings to consolidated warnings
                if stim_warnings:
                    if '_warnings' not in consolidated_data:
                        consolidated_data['_warnings'] = []
                    consolidated_data['_warnings'].extend(stim_warnings)

            if progress.wasCanceled():
                return

            progress.setLabelText("Saving Excel file and generating charts...")
            progress.setValue(85)
            QApplication.processEvents()

            self._save_consolidated_to_excel(consolidated_data, Path(save_path))
            progress.setValue(100)

            n_files = len(means_files) + len(breaths_files)

            # Check for warnings from consolidation
            if '_warnings' in consolidated_data:
                warnings_text = consolidated_data['_warnings']
                msg_box = QMessageBox(self.window)
                msg_box.setIcon(QMessageBox.Icon.Warning)
                msg_box.setWindowTitle("Consolidation Completed with Warnings")
                msg_box.setText(f"Consolidated {n_files} files successfully.\nSaved to: {save_path}\n\nHowever, some files required special handling:")
                msg_box.setDetailedText(warnings_text)
                msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
                msg_box.exec()
            else:
                QMessageBox.information(
                    self.window, "Success",
                    f"Consolidated {n_files} files.\nSaved to: {save_path}"
                )

        except Exception as e:
            progress.close()
            QMessageBox.critical(self.window, "Consolidation Error", f"Failed to consolidate:\n{e}")
            import traceback
            traceback.print_exc()
        finally:
            progress.close()

   


    def _consolidate_breaths_histograms(self, files: list[tuple[str, Path]]) -> dict:
        """
        Process breaths CSV files and create density histograms for each metric.
        Returns dict: {metric_name: DataFrame with histogram bins and densities per file}
        """
        import pandas as pd
        import numpy as np
        
        # Metrics to extract from breaths data
        metrics = ['if', 'amp_insp', 'amp_exp', 'area_insp', 'area_exp', 'ti', 'te', 'vent_proxy']
        
        # Regions in the breaths CSV
        regions = {
            'all': '',
            'baseline': '_baseline',
            'stim': '_stim', 
            'post': '_post'
        }
        
        consolidated = {}
        
        # Helper to calculate mean and SEM
        def calc_mean_sem(data_array):
            mean = np.nanmean(data_array, axis=1)
            n = np.sum(np.isfinite(data_array), axis=1)
            std = np.nanstd(data_array, axis=1, ddof=1)
            sem = np.where(n >= 2, std / np.sqrt(n), np.nan)
            return mean, sem
        
        # Process each metric (combine all regions into one sheet)
        for metric in metrics:
            combined_df = None
            
            # Process RAW data
            for region_name, suffix in regions.items():
                col_name = f"{metric}{suffix}" if suffix else metric
                
                # Collect data for this region
                all_data = []
                file_roots = []
                
                for root, path in files:
                    try:
                        df = pd.read_csv(path, low_memory=False)

                        if col_name in df.columns:
                            values = df[col_name].dropna().values
                            if len(values) > 0:
                                all_data.append(values)
                                file_roots.append(root)
                    except Exception as e:
                        print(f"Error reading {path}: {e}")
                
                if not all_data:
                    continue
                
                # Common bin edges for this region
                all_combined = np.concatenate(all_data)
                n_bins = 30
                bin_edges = np.histogram_bin_edges(all_combined, bins=n_bins)
                bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
                
                # Calculate density histogram for each file
                region_data = {f'bin_center_{region_name}': bin_centers}
                
                density_arrays = []
                for root, values in zip(file_roots, all_data):
                    counts, _ = np.histogram(values, bins=bin_edges)
                    # Convert to density (normalize so integral = 1)
                    bin_widths = np.diff(bin_edges)
                    density = counts / (counts.sum() * bin_widths)
                    region_data[f'{root}_{region_name}'] = density
                    density_arrays.append(density)
                
                # Calculate mean and SEM for this region
                if density_arrays:
                    density_matrix = np.column_stack(density_arrays)
                    mean, sem = calc_mean_sem(density_matrix)
                    region_data[f'mean_{region_name}'] = mean
                    region_data[f'sem_{region_name}'] = sem
                
                # Create DataFrame for this region
                region_df = pd.DataFrame(region_data)
                
                # Combine with other regions
                if combined_df is None:
                    combined_df = region_df
                else:
                    # Merge horizontally
                    combined_df = pd.concat([combined_df, region_df], axis=1)
                
                # Add blank column separator AFTER each region
                combined_df[''] = ''
            
            # Process NORMALIZED data
            for region_name, suffix in regions.items():
                col_name = f"{metric}{suffix}_norm" if suffix else f"{metric}_norm"
                
                # Collect normalized data for this region
                all_data_norm = []
                file_roots_norm = []
                
                for root, path in files:
                    try:
                        df = pd.read_csv(path, low_memory=False)

                        if col_name in df.columns:
                            values = df[col_name].dropna().values
                            if len(values) > 0:
                                all_data_norm.append(values)
                                file_roots_norm.append(root)
                    except Exception as e:
                        print(f"Error reading {path}: {e}")
                
                if not all_data_norm:
                    continue
                
                # Common bin edges for normalized region
                all_combined_norm = np.concatenate(all_data_norm)
                n_bins = 30
                bin_edges_norm = np.histogram_bin_edges(all_combined_norm, bins=n_bins)
                bin_centers_norm = (bin_edges_norm[:-1] + bin_edges_norm[1:]) / 2
                
                # Calculate density histogram for each file
                region_data_norm = {f'bin_center_{region_name}_norm': bin_centers_norm}
                
                density_arrays_norm = []
                for root, values in zip(file_roots_norm, all_data_norm):
                    counts, _ = np.histogram(values, bins=bin_edges_norm)
                    bin_widths = np.diff(bin_edges_norm)
                    density = counts / (counts.sum() * bin_widths)
                    region_data_norm[f'{root}_{region_name}_norm'] = density
                    density_arrays_norm.append(density)
                
                # Calculate mean and SEM for normalized region
                if density_arrays_norm:
                    density_matrix_norm = np.column_stack(density_arrays_norm)
                    mean_norm, sem_norm = calc_mean_sem(density_matrix_norm)
                    region_data_norm[f'mean_{region_name}_norm'] = mean_norm
                    region_data_norm[f'sem_{region_name}_norm'] = sem_norm
                
                # Create DataFrame for this normalized region
                region_df_norm = pd.DataFrame(region_data_norm)
                
                # Merge horizontally
                combined_df = pd.concat([combined_df, region_df_norm], axis=1)
                
                # Add blank column separator AFTER each normalized region
                combined_df[''] = ''

            # Process EUPNEA-NORMALIZED data
            for region_name, suffix in regions.items():
                col_name = f"{metric}{suffix}_norm_eupnea" if suffix else f"{metric}_norm_eupnea"

                # Collect eupnea-normalized data for this region
                all_data_eupnea = []
                file_roots_eupnea = []

                for root, path in files:
                    try:
                        df = pd.read_csv(path, low_memory=False)

                        if col_name in df.columns:
                            values = df[col_name].dropna().values
                            if len(values) > 0:
                                all_data_eupnea.append(values)
                                file_roots_eupnea.append(root)
                    except Exception as e:
                        print(f"Error reading {path}: {e}")

                if not all_data_eupnea:
                    continue

                # Common bin edges for eupnea-normalized region
                all_combined_eupnea = np.concatenate(all_data_eupnea)
                n_bins = 30
                bin_edges_eupnea = np.histogram_bin_edges(all_combined_eupnea, bins=n_bins)
                bin_centers_eupnea = (bin_edges_eupnea[:-1] + bin_edges_eupnea[1:]) / 2

                # Calculate density histogram for each file
                region_data_eupnea = {f'bin_center_{region_name}_eupnea': bin_centers_eupnea}

                density_arrays_eupnea = []
                for root, values in zip(file_roots_eupnea, all_data_eupnea):
                    counts, _ = np.histogram(values, bins=bin_edges_eupnea)
                    bin_widths = np.diff(bin_edges_eupnea)
                    density = counts / (counts.sum() * bin_widths)
                    region_data_eupnea[f'{root}_{region_name}_eupnea'] = density
                    density_arrays_eupnea.append(density)

                # Calculate mean and SEM for eupnea-normalized region
                if density_arrays_eupnea:
                    density_matrix_eupnea = np.column_stack(density_arrays_eupnea)
                    mean_eupnea, sem_eupnea = calc_mean_sem(density_matrix_eupnea)
                    region_data_eupnea[f'mean_{region_name}_eupnea'] = mean_eupnea
                    region_data_eupnea[f'sem_{region_name}_eupnea'] = sem_eupnea

                # Create DataFrame for this eupnea-normalized region
                region_df_eupnea = pd.DataFrame(region_data_eupnea)

                # Merge horizontally
                combined_df = pd.concat([combined_df, region_df_eupnea], axis=1)

                # Add blank column separator AFTER each eupnea-normalized region
                combined_df[''] = ''

            if combined_df is not None:
                consolidated[f'{metric}_histogram'] = {
                    'time_series': combined_df,
                    'raw_summary': {},
                    'norm_summary': {},
                    'eupnea_summary': {},
                    'windows': []
                }

        return consolidated



    def _consolidate_events(self, files: list[tuple[str, Path]]) -> pd.DataFrame:
        """
        Consolidate events CSV files from multiple experiments (excluding stimulus events).
        Adds experiment, experiment_number, and global_sweep_number columns.
        """
        import pandas as pd
        import numpy as np

        all_events = []

        for exp_num, (root, path) in enumerate(files, start=1):
            try:
                df = pd.read_csv(path, low_memory=False)

                if len(df) > 0:
                    # Filter out stimulus events
                    if 'event_type' in df.columns:
                        df = df[~df['event_type'].str.contains('stimulus', case=False, na=False)]

                    if len(df) > 0:  # Check if we still have data after filtering
                        # Add experiment identifier columns
                        df.insert(0, 'experiment', root)
                        df.insert(1, 'experiment_number', exp_num)

                        # Calculate global sweep number
                        # Each experiment has sweeps numbered 1, 2, 3, etc.
                        # We need to add an offset based on previous experiments
                        if 'sweep' in df.columns:
                            # Find the maximum sweep number in this experiment
                            max_sweep = df['sweep'].max()
                            # Calculate offset (total sweeps from previous experiments)
                            sweep_offset = sum([pd.read_csv(files[i][1], low_memory=False)['sweep'].max()
                                              for i in range(exp_num - 1)
                                              if 'sweep' in pd.read_csv(files[i][1], low_memory=False).columns])
                            df.insert(2, 'global_sweep_number', df['sweep'] + sweep_offset)

                        all_events.append(df)

            except Exception as e:
                print(f"Error reading events file {path}: {e}")

        if all_events:
            return pd.concat(all_events, ignore_index=True)
        else:
            return pd.DataFrame()


    def _consolidate_stimulus(self, files: list[tuple[str, Path]]) -> tuple[pd.DataFrame, list[str]]:
        """
        Extract stimulus events and validate consistency across all experiments/sweeps.
        Returns: (stimulus_df with one instance, list of warnings)
        """
        import pandas as pd
        import numpy as np

        all_stimulus = []
        warnings = []
        reference_stim = None

        for exp_num, (root, path) in enumerate(files, start=1):
            try:
                df = pd.read_csv(path, low_memory=False)

                if 'event_type' in df.columns:
                    # Extract only stimulus events
                    stim_events = df[df['event_type'].str.contains('stimulus', case=False, na=False)].copy()

                    if len(stim_events) > 0:
                        # Determine which column names are used (could be start_time/end_time or t_start/t_end)
                        start_col = 'start_time' if 'start_time' in stim_events.columns else 't_start'
                        end_col = 'end_time' if 'end_time' in stim_events.columns else 't_end'

                        # Get unique stimulus events (should be same across all sweeps)
                        unique_stim = stim_events.drop_duplicates(subset=['event_type', start_col, end_col])

                        if reference_stim is None:
                            # First experiment sets the reference
                            reference_stim = unique_stim[[start_col, end_col]].values
                            # Store with experiment identifier
                            result_df = unique_stim.copy()
                            result_df.insert(0, 'experiment', root)
                            all_stimulus.append(result_df)
                        else:
                            # Validate against reference
                            current_stim = unique_stim[[start_col, end_col]].values

                            # Check if stimulus events match
                            if not (len(reference_stim) == len(current_stim) and
                                    np.allclose(reference_stim.astype(float),
                                               current_stim.astype(float),
                                               rtol=1e-5)):
                                warnings.append(f"WARNING: Stimulus timing differs in experiment '{root}'")
                                print(f"  Stimulus mismatch in {root}")

            except Exception as e:
                print(f"Error reading stimulus from {path}: {e}")
                import traceback
                traceback.print_exc()

        if all_stimulus:
            # Return only the first instance (they should all be identical)
            return all_stimulus[0], warnings
        else:
            return pd.DataFrame(), warnings

    

    # -------------------- NPZ-based fast consolidation helpers --------------------


    def _try_load_npz_v2(self, npz_path: Path) -> dict | None:
        """
        Try to load NPZ version 2 bundle with full timeseries data.
        Returns dict with data if successful, None otherwise.
        """
        # Silently skip if file doesn't exist (expected for old format)
        if not npz_path.exists():
            return None

        try:
            data = np.load(npz_path, allow_pickle=True)

            # Check if this is version 2
            version = data.get('npz_version', None)
            if version is None or int(version) < 2:
                # Silently skip v1 NPZ files (will use CSV fallback)
                return None

            # Verify required keys exist
            required_keys = ['timeseries_t', 'timeseries_keys']
            if not all(k in data for k in required_keys):
                print(f"[NPZ] Warning: {npz_path.name} missing required keys, using CSV fallback")
                return None

            return dict(data)
        except Exception as e:
            # Only warn on actual load errors (not missing files)
            print(f"[NPZ] Warning: Failed to load {npz_path.name}: {e}")
            return None


    def _extract_timeseries_from_npz(self, npz_data: dict, metric: str, variant: str = 'raw') -> tuple[np.ndarray, np.ndarray]:
        """
        Extract timeseries data for a specific metric from NPZ bundle.

        Args:
            npz_data: Loaded NPZ data dictionary
            metric: Metric name (e.g., 'if', 'amp_insp')
            variant: One of 'raw', 'norm', 'eupnea'

        Returns:
            (t, Y) where t is time vector and Y is (M, S) metric matrix
        """
        t = npz_data['timeseries_t']

        if variant == 'raw':
            key = f'ts_raw_{metric}'
        elif variant == 'norm':
            key = f'ts_norm_{metric}'
        elif variant == 'eupnea':
            key = f'ts_eupnea_{metric}'
        else:
            raise ValueError(f"Unknown variant: {variant}")

        Y = npz_data[key]
        return t, Y


    def _consolidate_from_npz_v2(self, npz_data_by_root: dict, files: list[tuple[str, Path]], metrics: list[str]) -> dict:
        """
        Fast consolidation using NPZ v2 bundles with pre-computed timeseries data.
        Much faster than CSV because:
        1. No CSV parsing overhead
        2. No interpolation needed (all files share same time base from stimulus alignment)
        3. Direct numpy array operations
        """
        import pandas as pd
        import numpy as np

        # Helper function to calculate mean and SEM
        def calc_mean_sem(data_array):
            """Calculate mean and SEM from array of values."""
            n = np.sum(np.isfinite(data_array), axis=1)
            mean = np.full(data_array.shape[0], np.nan)
            sem = np.full(data_array.shape[0], np.nan)
            valid_rows = n > 0
            if valid_rows.any():
                mean[valid_rows] = np.nanmean(data_array[valid_rows, :], axis=1)
                sem_rows = n >= 2
                if sem_rows.any():
                    std = np.nanstd(data_array[sem_rows, :], axis=1, ddof=1)
                    sem[sem_rows] = std / np.sqrt(n[sem_rows])
            return mean, sem

        # Helper function to calculate window mean
        def window_mean(t, y, t_start, t_end):
            """Calculate mean of y values where t is between t_start and t_end."""
            mask = (t >= t_start) & (t < t_end)
            if mask.sum() == 0:
                return np.nan
            return np.nanmean(y[mask])

        # Define time windows for summary stats
        windows = [
            ('Baseline (-10 to 0s)', -10.0, 0.0),
            ('Baseline (-5 to 0s)', -5.0, 0.0),
            ('Stim (0-15s)', 0.0, 15.0),
            ('Stim (0-5s)', 0.0, 5.0),
            ('Stim (5-10s)', 5.0, 10.0),
            ('Stim (10-15s)', 10.0, 15.0),
            ('Post (15-25s)', 15.0, 25.0),
            ('Post (15-20s)', 15.0, 20.0),
            ('Post (20-25s)', 20.0, 25.0),
            ('Post (25-30s)', 25.0, 30.0),
        ]

        # Determine common time base by scanning all files
        # This allows experiments with different durations to be consolidated
        all_t_mins = []
        all_t_maxs = []
        all_steps = []

        for root, _ in files:
            t_file = npz_data_by_root[root]['timeseries_t']
            all_t_mins.append(t_file.min())
            all_t_maxs.append(t_file.max())
            if len(t_file) > 1:
                all_steps.append(np.median(np.diff(t_file)))

        # Create common time grid spanning the union of all time ranges
        t_common_min = min(all_t_mins)
        t_common_max = max(all_t_maxs)
        t_common_step = np.median(all_steps) if all_steps else 0.1

        # Generate uniform time grid
        t_common = np.arange(t_common_min, t_common_max + t_common_step/2, t_common_step)

        print(f"[NPZ] Common time grid: {t_common_min:.2f}s to {t_common_max:.2f}s, step={t_common_step:.4f}s ({len(t_common)} points)")

        consolidated = {}

        # Get first file root for checking metric existence
        first_root = files[0][0]

        # Process each metric
        for metric in metrics:
            # Check if metric exists in first NPZ
            test_key = f'ts_raw_{metric}'
            if test_key not in npz_data_by_root[first_root]:
                continue

            # Build all columns as dict first to avoid fragmentation
            data_dict = {'t': t_common}

            # Process each variant: raw, norm, eupnea
            # Note: eupnea suffix is '_eupnea' not '_norm_eupnea' to match CSV consolidation
            for variant, suffix in [('raw', ''), ('norm', '_norm'), ('eupnea', '_eupnea')]:
                # Collect data from all files
                all_means = []
                file_means = {}  # Store for window calculations

                for root, _ in files:
                    t_file, Y = self._extract_timeseries_from_npz(npz_data_by_root[root], metric, variant)

                    # Compute mean across sweeps (axis 1)
                    y_mean_file = np.nanmean(Y, axis=1)

                    # Interpolate to common time grid if needed
                    if len(t_file) != len(t_common) or not np.allclose(t_file, t_common, rtol=0.01):
                        from scipy.interpolate import interp1d
                        # Use linear interpolation, extrapolate with NaN outside range
                        interp_func = interp1d(t_file, y_mean_file, kind='linear',
                                              bounds_error=False, fill_value=np.nan)
                        y_mean = interp_func(t_common)
                    else:
                        y_mean = y_mean_file

                    all_means.append(y_mean)
                    file_means[root] = (t_file, y_mean_file)  # Store original for window calculations

                    # Add individual file column
                    data_dict[f'{root}{suffix}'] = y_mean

                # Stack all means into matrix (M, num_files)
                all_means_matrix = np.column_stack(all_means)

                # Compute mean and SEM across files
                mean, sem = calc_mean_sem(all_means_matrix)
                data_dict[f'mean{suffix}'] = mean
                data_dict[f'sem{suffix}'] = sem

                # Add separator column
                data_dict[f' {suffix}'] = ''

            # Create DataFrame once with all columns
            result_df = pd.DataFrame(data_dict)

            # Build summary data dicts (for Excel summary section)
            # These use the original (non-interpolated) data for window calculations
            # We need to collect these from each variant's file_means
            raw_summary = {}
            norm_summary = {}
            eupnea_summary = {}

            # Re-extract data for each variant to populate summary dicts
            for root, _ in files:
                # Raw data
                t_file, Y = self._extract_timeseries_from_npz(npz_data_by_root[root], metric, 'raw')
                y_mean_file = np.nanmean(Y, axis=1)
                raw_summary[root] = (t_file, y_mean_file)

                # Normalized data
                t_file, Y = self._extract_timeseries_from_npz(npz_data_by_root[root], metric, 'norm')
                y_mean_file = np.nanmean(Y, axis=1)
                norm_summary[root] = (t_file, y_mean_file)

                # Eupnea-normalized data
                t_file, Y = self._extract_timeseries_from_npz(npz_data_by_root[root], metric, 'eupnea')
                y_mean_file = np.nanmean(Y, axis=1)
                eupnea_summary[root] = (t_file, y_mean_file)

            # Package result in same format as CSV consolidation
            consolidated[metric] = {
                'time_series': result_df,
                'raw_summary': raw_summary,
                'norm_summary': norm_summary,
                'eupnea_summary': eupnea_summary,
                'windows': windows
            }

        print(f"[NPZ] Fast consolidation complete")
        return consolidated


    def _consolidate_means_files(self, files: list[tuple[str, Path]]) -> dict:
        """
        Consolidate means_by_time CSV files.
        Interpolates all data to a common time base and adds summary statistics.
        Returns dict: {metric_name: DataFrame with time series + summary stats}
        """
        import pandas as pd
        import numpy as np
        from scipy.interpolate import interp1d

        metrics = [
            'if', 'amp_insp', 'amp_exp', 'area_insp', 'area_exp',
            'ti', 'te', 'vent_proxy'
        ]

        # Try fast NPZ-based consolidation first (5-10× faster)
        npz_data_by_root = {}
        npz_success_count = 0
        for root, csv_path in files:
            # Look for NPZ bundle next to the CSV (handle both naming patterns)
            if csv_path.name.endswith('_timeseries.csv'):
                npz_name = csv_path.name.replace('_timeseries.csv', '_bundle.npz')
            elif csv_path.name.endswith('_means_by_time.csv'):
                npz_name = csv_path.name.replace('_means_by_time.csv', '_bundle.npz')
            else:
                continue  # Unknown CSV pattern

            npz_path = csv_path.parent / npz_name
            npz_data = self._try_load_npz_v2(npz_path)
            if npz_data is not None:
                npz_data_by_root[root] = npz_data
                npz_success_count += 1

        # If all files have NPZ v2, use fast path (with interpolation support)
        if npz_success_count == len(files) and npz_success_count > 0:
            print(f"[NPZ Fast Path] Loading {npz_success_count} files from NPZ bundles (v2)...")
            return self._consolidate_from_npz_v2(npz_data_by_root, files, metrics)
        elif npz_success_count > 0:
            print(f"[NPZ] Only {npz_success_count}/{len(files)} files have NPZ v2, falling back to CSV...")
        else:
            print(f"[CSV] No NPZ v2 bundles found, using CSV files...")

        # Determine common time base by scanning all files
        # This allows experiments with different durations to be consolidated
        print("Scanning all files to determine common time range...")
        all_t_mins = []
        all_t_maxs = []
        all_steps = []

        for root, path in files:
            df_temp = pd.read_csv(path, low_memory=False)
            t_temp = df_temp['t'].values
            all_t_mins.append(t_temp.min())
            all_t_maxs.append(t_temp.max())
            if len(t_temp) > 1:
                all_steps.append(np.median(np.diff(t_temp)))

        # Create common time grid spanning the union of all time ranges
        t_common_min = min(all_t_mins)
        t_common_max = max(all_t_maxs)
        t_common_step = np.median(all_steps) if all_steps else 0.1

        # Generate uniform time grid
        t_common = np.arange(t_common_min, t_common_max + t_common_step/2, t_common_step)

        print(f"Common time grid: {t_common_min:.2f}s to {t_common_max:.2f}s, step={t_common_step:.4f}s ({len(t_common)} points)")

        # Load first file for column checking
        first_root, first_path = files[0]
        df_first = pd.read_csv(first_path, low_memory=False)

        # Track files with potential issues
        warning_messages = []
        files_needing_interpolation = []
        files_with_poor_overlap = []
        files_with_different_sampling = []

        consolidated = {}
        
        # Helper function to calculate mean and SEM
        def calc_mean_sem(data_array):
            """Calculate mean and SEM from array of values."""
            # Count valid values per row
            n = np.sum(np.isfinite(data_array), axis=1)

            # Initialize outputs with NaN
            mean = np.full(data_array.shape[0], np.nan)
            sem = np.full(data_array.shape[0], np.nan)

            # Only calculate for rows with at least one valid value
            valid_rows = n > 0
            if valid_rows.any():
                mean[valid_rows] = np.nanmean(data_array[valid_rows, :], axis=1)

                # Only calculate SEM for rows with at least 2 values
                sem_rows = n >= 2
                if sem_rows.any():
                    std = np.nanstd(data_array[sem_rows, :], axis=1, ddof=1)
                    sem[sem_rows] = std / np.sqrt(n[sem_rows])

            return mean, sem
        
        # Helper function to calculate window mean
        def window_mean(t, y, t_start, t_end):
            """Calculate mean of y values where t is between t_start and t_end."""
            mask = (t >= t_start) & (t < t_end)
            if mask.sum() == 0:
                return np.nan
            return np.nanmean(y[mask])
        
        # Define time windows for summary stats
        windows = [
            ('Baseline (-10 to 0s)', -10.0, 0.0),
            ('Baseline (-5 to 0s)', -5.0, 0.0),
            ('Stim (0-15s)', 0.0, 15.0),
            ('Stim (0-5s)', 0.0, 5.0),
            ('Stim (5-10s)', 5.0, 10.0),
            ('Stim (10-15s)', 10.0, 15.0),
            ('Post (15-25s)', 15.0, 25.0),
            ('Post (15-20s)', 15.0, 20.0),
            ('Post (20-25s)', 20.0, 25.0),
            ('Post (25-30s)', 25.0, 30.0),
        ]
        
        # Process each metric (combining raw, time-normalized, and eupnea-normalized in same sheet)
        for metric in metrics:
            metric_mean_col = f"{metric}_mean"
            metric_norm_col = f"{metric}_norm_mean"
            metric_eupnea_col = f"{metric}_norm_eupnea_mean"

            # Check if metric exists
            if metric_mean_col not in df_first.columns:
                continue

            result_df = pd.DataFrame({'t': t_common})

            # Store data for summary calculations
            raw_data_dict = {}
            norm_data_dict = {}
            eupnea_data_dict = {}
            
            # Collect raw data from all files
            raw_data_cols = []
            for root, path in files:
                df = pd.read_csv(path, low_memory=False)

                if metric_mean_col not in df.columns:
                    print(f"Warning: {metric_mean_col} not found in {root}")
                    continue

                t_file = df['t'].values
                y_file = df[metric_mean_col].values

                # Check time range and sampling
                t_file_min, t_file_max = t_file.min(), t_file.max()
                t_file_step = np.median(np.diff(t_file)) if len(t_file) > 1 else np.nan

                # Calculate overlap percentage
                overlap_start = max(t_common_min, t_file_min)
                overlap_end = min(t_common_max, t_file_max)
                overlap_range = overlap_end - overlap_start
                common_range = t_common_max - t_common_min
                overlap_pct = 100 * overlap_range / common_range if common_range > 0 else 0

                # Check for different sampling rates
                if not np.isnan(t_file_step) and not np.isnan(t_common_step):
                    sampling_diff_pct = 100 * abs(t_file_step - t_common_step) / t_common_step
                    if sampling_diff_pct > 10 and root != first_root:
                        files_with_different_sampling.append(
                            f"{root}: {t_file_step:.4f}s vs reference {t_common_step:.4f}s ({sampling_diff_pct:.1f}% difference)"
                        )

                # Check for poor overlap
                if overlap_pct < 80 and root != first_root:
                    files_with_poor_overlap.append(
                        f"{root}: {overlap_pct:.1f}% overlap (range: {t_file_min:.1f} to {t_file_max:.1f}s)"
                    )

                # Always interpolate to common time grid (different files may have different durations)
                mask = np.isfinite(y_file)
                if mask.sum() >= 2:
                    try:
                        # Check if exact match first (optimization)
                        if len(t_file) == len(t_common) and np.allclose(t_file, t_common, rtol=1e-5, atol=1e-8):
                            result_df[root] = y_file
                            raw_data_dict[root] = (t_common, y_file)
                        else:
                            # Interpolate to common grid
                            if root not in files_needing_interpolation:
                                files_needing_interpolation.append(root)
                            print(f"Interpolating {root} to common time base for {metric}")
                            f_interp = interp1d(
                                t_file[mask], y_file[mask],
                                kind='linear',
                                bounds_error=False,
                                fill_value=np.nan
                            )
                            y_interp = f_interp(t_common)
                            result_df[root] = y_interp
                            raw_data_dict[root] = (t_common, y_interp)

                            # Count how many points were extrapolated (NaN after interpolation)
                            n_extrapolated = np.sum(np.isnan(y_interp) & ~np.isnan(t_common))
                            if n_extrapolated > 0:
                                extrap_pct = 100 * n_extrapolated / len(t_common)
                                if extrap_pct > 5:
                                    print(f"  Warning: {extrap_pct:.1f}% of points extrapolated (outside data range)")
                    except Exception as e:
                        print(f"Error interpolating {root} for {metric}: {e}")
                        result_df[root] = np.nan
                else:
                    print(f"  Warning: Insufficient data points for interpolation in {root}")
                    result_df[root] = np.nan

                raw_data_cols.append(root)
            
            # Calculate raw mean and SEM
            if raw_data_cols:
                raw_data = result_df[raw_data_cols].values
                result_df['mean'], result_df['sem'] = calc_mean_sem(raw_data)
            
            # Collect normalized data from all files
            if metric_norm_col in df_first.columns:
                norm_data_cols = []
                for root, path in files:
                    df = pd.read_csv(path, low_memory=False)
                    
                    if metric_norm_col not in df.columns:
                        continue
                    
                    t_file = df['t'].values
                    y_file = df[metric_norm_col].values
                    
                    norm_col_name = f"{root}_norm"

                    # Always interpolate to common time grid
                    mask = np.isfinite(y_file)
                    if mask.sum() >= 2:
                        try:
                            # Check if exact match first (optimization)
                            if len(t_file) == len(t_common) and np.allclose(t_file, t_common, rtol=1e-5, atol=1e-8):
                                result_df[norm_col_name] = y_file
                                norm_data_dict[root] = (t_common, y_file)
                            else:
                                f_interp = interp1d(
                                    t_file[mask], y_file[mask],
                                    kind='linear',
                                    bounds_error=False,
                                    fill_value=np.nan
                                )
                                y_interp = f_interp(t_common)
                                result_df[norm_col_name] = y_interp
                                norm_data_dict[root] = (t_common, y_interp)
                        except:
                            result_df[norm_col_name] = np.nan
                    else:
                        result_df[norm_col_name] = np.nan
                    
                    norm_data_cols.append(norm_col_name)
                
                # Calculate normalized mean and SEM
                if norm_data_cols:
                    norm_data = result_df[norm_data_cols].values
                    result_df['mean_norm'], result_df['sem_norm'] = calc_mean_sem(norm_data)

            # Collect eupnea-normalized data from all files
            if metric_eupnea_col in df_first.columns:
                eupnea_data_cols = []
                for root, path in files:
                    df = pd.read_csv(path, low_memory=False)

                    if metric_eupnea_col not in df.columns:
                        continue

                    t_file = df['t'].values
                    y_file = df[metric_eupnea_col].values

                    eupnea_col_name = f"{root}_eupnea"

                    # Always interpolate to common time grid
                    mask = np.isfinite(y_file)
                    if mask.sum() >= 2:
                        try:
                            # Check if exact match first (optimization)
                            if len(t_file) == len(t_common) and np.allclose(t_file, t_common, rtol=1e-5, atol=1e-8):
                                result_df[eupnea_col_name] = y_file
                                eupnea_data_dict[root] = (t_common, y_file)
                            else:
                                f_interp = interp1d(
                                    t_file[mask], y_file[mask],
                                    kind='linear',
                                    bounds_error=False,
                                    fill_value=np.nan
                                )
                                y_interp = f_interp(t_common)
                                result_df[eupnea_col_name] = y_interp
                                eupnea_data_dict[root] = (t_common, y_interp)
                        except:
                            result_df[eupnea_col_name] = np.nan
                    else:
                        result_df[eupnea_col_name] = np.nan

                    eupnea_data_cols.append(eupnea_col_name)

                # Calculate eupnea-normalized mean and SEM
                if eupnea_data_cols:
                    eupnea_data = result_df[eupnea_data_cols].values
                    result_df['mean_eupnea'], result_df['sem_eupnea'] = calc_mean_sem(eupnea_data)

            # Insert blank columns between data blocks
            # First blank: before time-normalized data
            norm_start_idx = None
            for i, col in enumerate(result_df.columns):
                if '_norm' in str(col) and '_eupnea' not in str(col):
                    norm_start_idx = i
                    break
            if norm_start_idx is not None:
                result_df.insert(norm_start_idx, '', '')

            # Second blank: before eupnea-normalized data
            eupnea_start_idx = None
            for i, col in enumerate(result_df.columns):
                if '_eupnea' in str(col):
                    eupnea_start_idx = i
                    break
            if eupnea_start_idx is not None:
                result_df.insert(eupnea_start_idx, ' ', '')

            # Build summary statistics (as rows below the time series)
            # This will be saved as a separate section in Excel
            consolidated[metric] = {
                'time_series': result_df,
                'raw_summary': raw_data_dict,
                'norm_summary': norm_data_dict,
                'eupnea_summary': eupnea_data_dict,
                'windows': windows
            }

        # Build warning summary
        if files_needing_interpolation or files_with_poor_overlap or files_with_different_sampling:
            warning_parts = []

            if files_needing_interpolation:
                warning_parts.append("FILES REQUIRING INTERPOLATION:")
                warning_parts.append(f"Reference file (no interpolation): {first_root}")
                warning_parts.append(f"Time range: {t_common_min:.2f} to {t_common_max:.2f}s")
                warning_parts.append(f"Sample interval: {t_common_step:.4f}s\n")
                for f in files_needing_interpolation:
                    warning_parts.append(f"  • {f}")
                warning_parts.append("")

            if files_with_different_sampling:
                warning_parts.append("FILES WITH DIFFERENT SAMPLING RATES:")
                for msg in files_with_different_sampling:
                    warning_parts.append(f"  • {msg}")
                warning_parts.append("")

            if files_with_poor_overlap:
                warning_parts.append("FILES WITH POOR TIME OVERLAP (<80%):")
                for msg in files_with_poor_overlap:
                    warning_parts.append(f"  • {msg}")
                warning_parts.append("")

            # Store warning message for display after processing
            consolidated['_warnings'] = '\n'.join(warning_parts)

        return consolidated



    def _consolidate_breaths_sighs(self, files: list[tuple[str, Path]]) -> pd.DataFrame:
        """
        Extract all breaths marked as sighs (is_sigh == 1) from breaths CSV files.
        Adds experiment_number and global_sweep_number columns.
        Returns DataFrame with sigh breaths from all files.
        """
        import pandas as pd
        import numpy as np

        # Columns to extract for raw data
        raw_cols = ['sweep', 'breath', 't', 'region', 'is_sigh',
                    'if', 'amp_insp', 'amp_exp', 'area_insp', 'area_exp',
                    'ti', 'te', 'vent_proxy']

        # Columns to extract for normalized data
        norm_cols = ['sweep', 'breath', 't', 'region', 'is_sigh',
                    'if_norm', 'amp_insp_norm', 'amp_exp_norm', 'area_insp_norm', 'area_exp_norm',
                    'ti_norm', 'te_norm', 'vent_proxy_norm']

        all_sighs_raw = []
        all_sighs_norm = []

        for exp_num, (root, path) in enumerate(files, start=1):
            try:
                df = pd.read_csv(path, low_memory=False)

                # Filter for sighs (is_sigh == 1)
                if 'is_sigh' in df.columns:
                    sigh_mask = df['is_sigh'] == 1

                    # Extract raw sigh data
                    available_raw_cols = [col for col in raw_cols if col in df.columns]
                    if available_raw_cols and sigh_mask.sum() > 0:
                        sigh_df_raw = df.loc[sigh_mask, available_raw_cols].copy()
                        sigh_df_raw.insert(0, 'file', root)
                        sigh_df_raw.insert(1, 'experiment_number', exp_num)

                        # Calculate global sweep number
                        if 'sweep' in sigh_df_raw.columns:
                            # Calculate offset from previous experiments
                            sweep_offset = sum([pd.read_csv(files[i][1], low_memory=False)['sweep'].max()
                                              for i in range(exp_num - 1)
                                              if 'sweep' in pd.read_csv(files[i][1], low_memory=False).columns])
                            sigh_df_raw.insert(2, 'global_sweep_number', sigh_df_raw['sweep'] + sweep_offset)

                        all_sighs_raw.append(sigh_df_raw)

                    # Extract normalized sigh data
                    available_norm_cols = [col for col in norm_cols if col in df.columns]
                    if available_norm_cols and sigh_mask.sum() > 0:
                        sigh_df_norm = df.loc[sigh_mask, available_norm_cols].copy()
                        sigh_df_norm.insert(0, 'file', root)
                        all_sighs_norm.append(sigh_df_norm)

            except Exception as e:
                print(f"Error reading sighs from {path}: {e}")

        # Combine all sigh data
        if all_sighs_raw:
            combined_raw = pd.concat(all_sighs_raw, ignore_index=True)
        else:
            combined_raw = pd.DataFrame(columns=['file', 'experiment_number', 'global_sweep_number'] + raw_cols)

        if all_sighs_norm:
            combined_norm = pd.concat(all_sighs_norm, ignore_index=True)
        else:
            combined_norm = pd.DataFrame(columns=['file'] + norm_cols)

        # Combine raw and normalized with blank column separator
        combined_raw[''] = ''
        combined_sighs = pd.concat([combined_raw, combined_norm], axis=1)

        return combined_sighs

   


    def _save_consolidated_to_excel(self, consolidated: dict, save_path: Path):
        """Save consolidated dataframes to a single Excel file with multiple sheets."""
        import pandas as pd
        import numpy as np
        from openpyxl import load_workbook
        from openpyxl.styles import Font
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.chart import ScatterChart, Reference, Series
        from openpyxl.chart.marker import Marker  # Fixed typo here
        
        # Helper to calculate window means
        def window_mean(t, y, t_start, t_end):
            mask = (t >= t_start) & (t < t_end)
            if mask.sum() == 0:
                return np.nan
            return np.nanmean(y[mask])
        
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            # Define sheet order: time series metrics first, then sighs, then events/stimulus, then histograms
            time_series_metrics = []
            sighs_sheets = []
            event_sheets = []
            histogram_sheets = []

            for metric_name in consolidated.keys():
                if metric_name == '_warnings':
                    continue
                elif metric_name == 'sighs':
                    sighs_sheets.append(metric_name)
                elif metric_name in ['events', 'stimulus']:
                    event_sheets.append(metric_name)
                elif '_histogram' in metric_name:
                    histogram_sheets.append(metric_name)
                else:
                    time_series_metrics.append(metric_name)

            # Process sheets in desired order: time series -> sighs -> events/stimulus -> histograms
            ordered_sheets = time_series_metrics + sighs_sheets + event_sheets + histogram_sheets

            for metric_name in ordered_sheets:
                data_dict = consolidated[metric_name]
                time_series_df = data_dict['time_series']
                raw_summary = data_dict.get('raw_summary', {})
                norm_summary = data_dict.get('norm_summary', {})
                eupnea_summary = data_dict.get('eupnea_summary', {})
                windows = data_dict.get('windows', [])
                
                sheet_name = metric_name[:31]
                
                # Write time series data starting at A1
                time_series_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)
                
                # Only add summary if it exists (means files have it, histogram files don't)
                if raw_summary:
                    # Calculate starting column for summary (after time series + 2 blank columns)
                    summary_start_col = len(time_series_df.columns) + 2
                    
                    # Get the worksheet to write summary data
                    worksheet = writer.sheets[sheet_name]
                    
                    # Build raw summary DataFrame
                    raw_summary_rows = []
                    for root in raw_summary.keys():
                        t, y = raw_summary[root]
                        row = {'File': root}
                        for window_name, t_start, t_end in windows:
                            row[window_name] = window_mean(t, y, t_start, t_end)
                        raw_summary_rows.append(row)
                    
                    if raw_summary_rows:
                        raw_summary_df = pd.DataFrame(raw_summary_rows)
                        
                        # Write raw summary starting at top right
                        for r_idx, row in enumerate(dataframe_to_rows(raw_summary_df, index=False, header=True)):
                            for c_idx, value in enumerate(row):
                                worksheet.cell(row=r_idx + 1, column=summary_start_col + c_idx, value=value)
                    
                    # Build normalized summary DataFrame  
                    if norm_summary:
                        # Start normalized summary after raw summary + 2 blank columns
                        norm_start_col = summary_start_col + len(raw_summary_df.columns) + 2
                        
                        norm_summary_rows = []
                        for root in norm_summary.keys():
                            t, y = norm_summary[root]
                            row = {'File': root}
                            for window_name, t_start, t_end in windows:
                                row[f"{window_name}_norm"] = window_mean(t, y, t_start, t_end)
                            norm_summary_rows.append(row)
                        
                        if norm_summary_rows:
                            norm_summary_df = pd.DataFrame(norm_summary_rows)

                            # Write normalized summary to the right of raw summary
                            for r_idx, row in enumerate(dataframe_to_rows(norm_summary_df, index=False, header=True)):
                                for c_idx, value in enumerate(row):
                                    worksheet.cell(row=r_idx + 1, column=norm_start_col + c_idx, value=value)

                    # Build eupnea-normalized summary DataFrame
                    if eupnea_summary:
                        # Start eupnea summary after normalized summary + 2 blank columns
                        eupnea_start_col = norm_start_col + len(norm_summary_df.columns) + 2 if norm_summary else summary_start_col

                        eupnea_summary_rows = []
                        for root in eupnea_summary.keys():
                            t, y = eupnea_summary[root]
                            row = {'File': root}
                            for window_name, t_start, t_end in windows:
                                row[f"{window_name}_eupnea"] = window_mean(t, y, t_start, t_end)
                            eupnea_summary_rows.append(row)

                        if eupnea_summary_rows:
                            eupnea_summary_df = pd.DataFrame(eupnea_summary_rows)

                            # Write eupnea summary to the right of normalized summary
                            for r_idx, row in enumerate(dataframe_to_rows(eupnea_summary_df, index=False, header=True)):
                                for c_idx, value in enumerate(row):
                                    worksheet.cell(row=r_idx + 1, column=eupnea_start_col + c_idx, value=value)

                print(f"Saved sheet: {sheet_name}")
        
        # Apply bold formatting and add charts
        wb = load_workbook(save_path)
        bold_font = Font(bold=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"Processing sheet: '{sheet_name}'")

            # Bold header row only (much faster than bolding entire columns)
            header_row = ws[1]

            for cell in header_row:
                cell.font = bold_font
            
            # Add charts based on sheet type
            if sheet_name == 'sighs':
                print(f"Matched sighs sheet condition for '{sheet_name}'")
                # Create sigh timeline chart
                self._add_sighs_chart(ws, header_row)

            elif sheet_name == 'events':
                print(f"Matched events sheet condition for '{sheet_name}'")
                # Create eupnea and apnea timeline charts
                self._add_events_charts(ws, header_row)

            # Stimulus sheet - no charts needed (just tabular data)
            elif sheet_name == 'stimulus':
                pass  # No charts for stimulus

            # Add charts for histogram sheets
            elif '_histogram' in sheet_name:
                regions = ['all', 'baseline', 'stim', 'post']
                
                # Chart 1: Raw means overlay (reverted to original style)
                chart1 = ScatterChart()
                chart1.title = f"{sheet_name} - Raw Mean Histograms"
                chart1.style = 2
                chart1.x_axis.title = "Bin Center"
                chart1.y_axis.title = "Density"

                # Enable axes display
                chart1.x_axis.delete = False
                chart1.y_axis.delete = False

                # Enable axis tick marks and labels (major only, no minor)
                chart1.x_axis.tickLblPos = "nextTo"  # Changed from "low" to fix label positioning
                chart1.y_axis.tickLblPos = "nextTo"
                chart1.x_axis.majorTickMark = "out"
                chart1.y_axis.majorTickMark = "out"
                chart1.x_axis.minorTickMark = "none"
                chart1.y_axis.minorTickMark = "none"

                # Disable gridlines
                chart1.x_axis.majorGridlines = None
                chart1.y_axis.majorGridlines = None

                # Set y-axis to start at 0
                chart1.y_axis.scaling.min = 0
                
                for region in regions:
                    bin_col = None
                    mean_col = None
                    
                    for idx, cell in enumerate(header_row, start=1):
                        if cell.value == f'bin_center_{region}':
                            bin_col = idx
                        elif cell.value == f'mean_{region}':
                            mean_col = idx
                    
                    if bin_col and mean_col:
                        xvalues = Reference(ws, min_col=bin_col, min_row=2, max_row=ws.max_row)
                        yvalues = Reference(ws, min_col=mean_col, min_row=2, max_row=ws.max_row)
                        
                        series = Series(yvalues, xvalues, title=region)
                        chart1.series.append(series)
                
                chart1.width = 10
                chart1.height = 6
                ws.add_chart(chart1, f"A{ws.max_row + 3}")

                # Chart 2: Time-normalized means overlay
                chart2 = ScatterChart()
                chart2.title = f"{sheet_name} - Time-Normalized Mean Histograms"
                chart2.style = 2
                chart2.x_axis.title = "Bin Center (time-normalized)"
                chart2.y_axis.title = "Density"

                # Enable axes display
                chart2.x_axis.delete = False
                chart2.y_axis.delete = False

                # Enable axis tick marks and labels (major only, no minor)
                chart2.x_axis.tickLblPos = "nextTo"
                chart2.y_axis.tickLblPos = "nextTo"
                chart2.x_axis.majorTickMark = "out"
                chart2.y_axis.majorTickMark = "out"
                chart2.x_axis.minorTickMark = "none"
                chart2.y_axis.minorTickMark = "none"

                # Disable gridlines
                chart2.x_axis.majorGridlines = None
                chart2.y_axis.majorGridlines = None

                # Set y-axis to start at 0
                chart2.y_axis.scaling.min = 0

                for region in regions:
                    bin_col_norm = None
                    mean_col_norm = None

                    for idx, cell in enumerate(header_row, start=1):
                        if cell.value == f'bin_center_{region}_norm':
                            bin_col_norm = idx
                        elif cell.value == f'mean_{region}_norm':
                            mean_col_norm = idx

                    if bin_col_norm and mean_col_norm:
                        xvalues = Reference(ws, min_col=bin_col_norm, min_row=2, max_row=ws.max_row)
                        yvalues = Reference(ws, min_col=mean_col_norm, min_row=2, max_row=ws.max_row)

                        series = Series(yvalues, xvalues, title=f"{region}_norm")
                        chart2.series.append(series)

                chart2.width = 10
                chart2.height = 6
                ws.add_chart(chart2, f"K{ws.max_row + 3}")

                # Chart 3: Eupnea-normalized means overlay
                chart3 = ScatterChart()
                chart3.title = f"{sheet_name} - Eupnea-Normalized Mean Histograms"
                chart3.style = 2
                chart3.x_axis.title = "Bin Center (eupnea-normalized)"
                chart3.y_axis.title = "Density"

                # Enable axes display
                chart3.x_axis.delete = False
                chart3.y_axis.delete = False

                # Enable axis tick marks and labels (major only, no minor)
                chart3.x_axis.tickLblPos = "nextTo"
                chart3.y_axis.tickLblPos = "nextTo"
                chart3.x_axis.majorTickMark = "out"
                chart3.y_axis.majorTickMark = "out"
                chart3.x_axis.minorTickMark = "none"
                chart3.y_axis.minorTickMark = "none"

                # Disable gridlines
                chart3.x_axis.majorGridlines = None
                chart3.y_axis.majorGridlines = None

                # Set y-axis to start at 0
                chart3.y_axis.scaling.min = 0

                for region in regions:
                    bin_col_eupnea = None
                    mean_col_eupnea = None

                    for idx, cell in enumerate(header_row, start=1):
                        if cell.value == f'bin_center_{region}_eupnea':
                            bin_col_eupnea = idx
                        elif cell.value == f'mean_{region}_eupnea':
                            mean_col_eupnea = idx

                    if bin_col_eupnea and mean_col_eupnea:
                        xvalues = Reference(ws, min_col=bin_col_eupnea, min_row=2, max_row=ws.max_row)
                        yvalues = Reference(ws, min_col=mean_col_eupnea, min_row=2, max_row=ws.max_row)

                        series = Series(yvalues, xvalues, title=f"{region}_eupnea")
                        chart3.series.append(series)

                chart3.width = 10
                chart3.height = 6
                ws.add_chart(chart3, f"U{ws.max_row + 3}")  # Position to the right of chart2

            # Add charts for time series sheets (not histograms)
            elif '_histogram' not in sheet_name:
                t_col = None
                mean_col = None
                sem_col = None
                mean_norm_col = None
                sem_norm_col = None
                mean_eupnea_col = None
                sem_eupnea_col = None

                # Find t, mean, sem, mean_norm, sem_norm, mean_eupnea, and sem_eupnea columns
                for idx, cell in enumerate(header_row, start=1):
                    if cell.value == 't':
                        t_col = idx
                    elif cell.value == 'mean':
                        mean_col = idx
                    elif cell.value == 'sem':
                        sem_col = idx
                    elif cell.value == 'mean_norm':
                        mean_norm_col = idx
                    elif cell.value == 'sem_norm':
                        sem_norm_col = idx
                    elif cell.value == 'mean_eupnea':
                        mean_eupnea_col = idx
                    elif cell.value == 'sem_eupnea':
                        sem_eupnea_col = idx
                
                # Position charts near top of sheet (row 5)
                chart_row = 5

                # Find individual file columns for plotting
                # Raw data columns: between 'mean' and first '_norm' column
                raw_file_cols = []
                norm_file_cols = []
                eupnea_file_cols = []

                for idx, cell in enumerate(header_row, start=1):
                    col_name = str(cell.value or '')
                    # Skip time, mean, sem, and blank columns
                    if col_name in ['t', 'mean', 'sem', 'mean_norm', 'sem_norm', 'mean_eupnea', 'sem_eupnea', '', 'None']:
                        continue
                    # Skip columns with numeric or empty names
                    if not col_name or col_name.isspace():
                        continue
                    if '_eupnea' in col_name:
                        eupnea_file_cols.append(idx)
                    elif '_norm' in col_name:
                        norm_file_cols.append(idx)
                    else:
                        # Must be a raw data file column
                        if t_col and idx > t_col:  # Make sure it's after the time column
                            raw_file_cols.append(idx)

                # Chart 1: Raw mean vs time
                if t_col and mean_col:
                    chart1 = ScatterChart()
                    chart1.title = f"{sheet_name} - Mean vs Time (Raw)"
                    chart1.style = 13

                    chart1.x_axis.title = "Time (s)"
                    chart1.y_axis.title = sheet_name

                    # Enable axes display
                    chart1.x_axis.delete = False
                    chart1.y_axis.delete = False

                    # Enable axis tick marks and labels (major only, no minor)
                    chart1.x_axis.tickLblPos = "nextTo"
                    chart1.y_axis.tickLblPos = "nextTo"
                    chart1.x_axis.majorTickMark = "out"
                    chart1.y_axis.majorTickMark = "out"
                    chart1.x_axis.minorTickMark = "none"
                    chart1.y_axis.minorTickMark = "none"

                    # Position y-axis on the left side
                    chart1.y_axis.crosses = "min"

                    # Disable gridlines
                    chart1.x_axis.majorGridlines = None
                    chart1.y_axis.majorGridlines = None

                    # Hide legend
                    chart1.legend = None

                    xvalues = Reference(ws, min_col=t_col, min_row=2, max_row=ws.max_row)

                    # Add individual file traces in light gray
                    for file_col in raw_file_cols:
                        yvalues_file = Reference(ws, min_col=file_col, min_row=2, max_row=ws.max_row)
                        series_file = Series(yvalues_file, xvalues, title="")
                        series_file.marker = Marker('none')
                        series_file.smooth = True
                        series_file.graphicalProperties.line.solidFill = "D3D3D3"  # Light gray
                        series_file.graphicalProperties.line.width = 8000  # Thinner than mean
                        chart1.series.append(series_file)

                    # Add mean line on top
                    yvalues = Reference(ws, min_col=mean_col, min_row=2, max_row=ws.max_row)
                    series = Series(yvalues, xvalues, title="Mean")
                    series.marker = Marker('none')
                    series.smooth = True
                    series.graphicalProperties.line.solidFill = "4472C4"  # Solid blue line
                    series.graphicalProperties.line.width = 12700  # 1pt line width
                    chart1.series.append(series)

                    chart1.width = 10
                    chart1.height = 6
                    ws.add_chart(chart1, f"A{chart_row}")
                
                # Chart 2: Time-normalized mean vs time
                if t_col and mean_norm_col:
                    chart2 = ScatterChart()
                    chart2.title = f"{sheet_name} - Mean vs Time (Time-Normalized)"
                    chart2.style = 13

                    chart2.x_axis.title = "Time (s)"
                    chart2.y_axis.title = f"{sheet_name} (time-normalized)"

                    # Enable axes display
                    chart2.x_axis.delete = False
                    chart2.y_axis.delete = False

                    # Enable axis tick marks and labels (major only, no minor)
                    chart2.x_axis.tickLblPos = "nextTo"
                    chart2.y_axis.tickLblPos = "nextTo"
                    chart2.x_axis.majorTickMark = "out"
                    chart2.y_axis.majorTickMark = "out"
                    chart2.x_axis.minorTickMark = "none"
                    chart2.y_axis.minorTickMark = "none"

                    # Position y-axis on the left side
                    chart2.y_axis.crosses = "min"

                    # Disable gridlines
                    chart2.x_axis.majorGridlines = None
                    chart2.y_axis.majorGridlines = None

                    # Hide legend
                    chart2.legend = None

                    xvalues = Reference(ws, min_col=t_col, min_row=2, max_row=ws.max_row)

                    # Add individual file traces in light gray
                    for file_col in norm_file_cols:
                        yvalues_file = Reference(ws, min_col=file_col, min_row=2, max_row=ws.max_row)
                        series_file = Series(yvalues_file, xvalues, title="")
                        series_file.marker = Marker('none')
                        series_file.smooth = True
                        series_file.graphicalProperties.line.solidFill = "D3D3D3"  # Light gray
                        series_file.graphicalProperties.line.width = 8000  # Thinner than mean
                        chart2.series.append(series_file)

                    # Add mean line on top
                    yvalues = Reference(ws, min_col=mean_norm_col, min_row=2, max_row=ws.max_row)
                    series = Series(yvalues, xvalues, title="Mean")
                    series.marker = Marker('none')
                    series.smooth = True
                    series.graphicalProperties.line.solidFill = "ED7D31"  # Solid orange line
                    series.graphicalProperties.line.width = 12700  # 1pt line width
                    chart2.series.append(series)

                    chart2.width = 10
                    chart2.height = 6
                    ws.add_chart(chart2, f"K{chart_row}")

                # Chart 3: Eupnea-normalized mean vs time
                if t_col and mean_eupnea_col:
                    chart3 = ScatterChart()
                    chart3.title = f"{sheet_name} - Mean vs Time (Eupnea-Normalized)"
                    chart3.style = 13

                    chart3.x_axis.title = "Time (s)"
                    chart3.y_axis.title = f"{sheet_name} (eupnea-normalized)"

                    # Enable axes display
                    chart3.x_axis.delete = False
                    chart3.y_axis.delete = False

                    # Enable axis tick marks and labels (major only, no minor)
                    chart3.x_axis.tickLblPos = "nextTo"
                    chart3.y_axis.tickLblPos = "nextTo"
                    chart3.x_axis.majorTickMark = "out"
                    chart3.y_axis.majorTickMark = "out"
                    chart3.x_axis.minorTickMark = "none"
                    chart3.y_axis.minorTickMark = "none"

                    # Position y-axis on the left side
                    chart3.y_axis.crosses = "min"

                    # Disable gridlines
                    chart3.x_axis.majorGridlines = None
                    chart3.y_axis.majorGridlines = None

                    # Hide legend
                    chart3.legend = None

                    xvalues = Reference(ws, min_col=t_col, min_row=2, max_row=ws.max_row)

                    # Add individual file traces in light gray
                    for file_col in eupnea_file_cols:
                        yvalues_file = Reference(ws, min_col=file_col, min_row=2, max_row=ws.max_row)
                        series_file = Series(yvalues_file, xvalues, title="")
                        series_file.marker = Marker('none')
                        series_file.smooth = True
                        series_file.graphicalProperties.line.solidFill = "D3D3D3"  # Light gray
                        series_file.graphicalProperties.line.width = 8000  # Thinner than mean
                        chart3.series.append(series_file)

                    # Add mean line on top
                    yvalues = Reference(ws, min_col=mean_eupnea_col, min_row=2, max_row=ws.max_row)
                    series = Series(yvalues, xvalues, title="Mean")
                    series.marker = Marker('none')
                    series.smooth = True
                    series.graphicalProperties.line.solidFill = "70AD47"  # Solid green line
                    series.graphicalProperties.line.width = 12700  # 1pt line width
                    chart3.series.append(series)

                    chart3.width = 10
                    chart3.height = 6
                    ws.add_chart(chart3, f"U{chart_row}")  # Position to the right of chart2

        wb.save(save_path)
        print(f"Applied bold formatting and charts. Consolidated Excel file saved: {save_path}")



    def _add_events_charts(self, ws, header_row):
        """Add eupnea, apnea, and sniffing timeline charts to the events sheet."""
        from openpyxl.chart import ScatterChart, Reference, Series
        from openpyxl.chart.marker import Marker
        import matplotlib.pyplot as plt
        import matplotlib.patches as patches
        from matplotlib.backends.backend_agg import FigureCanvasAgg
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage
        import io
        import numpy as np

        print(f"_add_events_charts called for sheet with {ws.max_row} rows")

        # Find required columns (handle both naming conventions)
        exp_num_col = None
        sweep_col = None
        global_sweep_col = None
        event_type_col = None
        t_start_col = None
        t_end_col = None

        for idx, cell in enumerate(header_row, start=1):
            if cell.value == 'experiment_number':
                exp_num_col = idx
            elif cell.value == 'sweep':
                sweep_col = idx
            elif cell.value == 'global_sweep_number':
                global_sweep_col = idx
            elif cell.value == 'event_type':
                event_type_col = idx
            elif cell.value in ['t_start', 'start_time']:
                t_start_col = idx
            elif cell.value in ['t_end', 'end_time']:
                t_end_col = idx

        if not all([exp_num_col, global_sweep_col, event_type_col, t_start_col, t_end_col]):
            print(f"Events sheet missing required columns for charts. Found columns: {[cell.value for cell in header_row]}")
            return

        # Read data from sheet
        events = []
        for row_idx in range(2, ws.max_row + 1):
            exp_num = ws.cell(row=row_idx, column=exp_num_col).value
            global_sweep = ws.cell(row=row_idx, column=global_sweep_col).value
            event_type = ws.cell(row=row_idx, column=event_type_col).value
            t_start = ws.cell(row=row_idx, column=t_start_col).value
            t_end = ws.cell(row=row_idx, column=t_end_col).value

            if all([exp_num is not None, global_sweep is not None,
                    event_type is not None, t_start is not None, t_end is not None]):
                events.append({
                    'exp_num': int(exp_num),
                    'global_sweep': int(global_sweep),
                    'event_type': str(event_type).lower(),
                    't_start': float(t_start),
                    't_end': float(t_end)
                })

        if not events:
            print("No events found for chart generation")
            return

        # Separate eupnea, apnea, and sniffing events
        eupnea_events = [e for e in events if 'eupnea' in e['event_type']]
        apnea_events = [e for e in events if 'apnea' in e['event_type']]
        sniffing_events = [e for e in events if 'sniff' in e['event_type']]

        if not eupnea_events and not apnea_events and not sniffing_events:
            print("No eupnea, apnea, or sniffing events found for charts")
            return

        # Create colormap for experiments
        n_experiments = max([e['exp_num'] for e in events])
        colors = plt.cm.tab10(np.linspace(0, 1, n_experiments))

        # Position charts at the top (starting at row 1)
        # We'll place them to the right of the data columns
        chart_start_row = 1

        # Create eupnea chart
        if eupnea_events:
            fig1, ax1 = plt.subplots(figsize=(10, 6))
            for event in eupnea_events:
                color = colors[event['exp_num'] - 1]
                ax1.plot([event['t_start'], event['t_end']],
                        [event['global_sweep'], event['global_sweep']],
                        color=color, linewidth=2, solid_capstyle='butt')

            ax1.set_xlabel('Time (s)')
            ax1.set_ylabel('Global Sweep Number')
            ax1.set_title('Eupnea Periods Across Experiments')
            ax1.grid(True, alpha=0.3)

            # Save figure to bytes
            buf1 = io.BytesIO()
            fig1.savefig(buf1, format='png', dpi=100, bbox_inches='tight')
            buf1.seek(0)
            plt.close(fig1)

            # Insert image into Excel
            img1 = XLImage(buf1)
            img1.width = 600
            img1.height = 360
            ws.add_image(img1, f"A{chart_start_row}")
            print(f"Added eupnea chart at A{chart_start_row}")

        # Create apnea chart
        if apnea_events:
            fig2, ax2 = plt.subplots(figsize=(10, 6))
            for event in apnea_events:
                color = colors[event['exp_num'] - 1]
                ax2.plot([event['t_start'], event['t_end']],
                        [event['global_sweep'], event['global_sweep']],
                        color=color, linewidth=2, solid_capstyle='butt')

            ax2.set_xlabel('Time (s)')
            ax2.set_ylabel('Global Sweep Number')
            ax2.set_title('Apnea Periods Across Experiments')
            ax2.grid(True, alpha=0.3)

            # Save figure to bytes
            buf2 = io.BytesIO()
            fig2.savefig(buf2, format='png', dpi=100, bbox_inches='tight')
            buf2.seek(0)
            plt.close(fig2)

            # Insert image into Excel
            img2 = XLImage(buf2)
            img2.width = 600
            img2.height = 360
            # Position to the right of first chart (column K)
            ws.add_image(img2, f"K{chart_start_row}")
            print(f"Added apnea chart at K{chart_start_row}")

        # Create sniffing chart
        if sniffing_events:
            fig3, ax3 = plt.subplots(figsize=(10, 6))
            for event in sniffing_events:
                color = colors[event['exp_num'] - 1]
                ax3.plot([event['t_start'], event['t_end']],
                        [event['global_sweep'], event['global_sweep']],
                        color=color, linewidth=2, solid_capstyle='butt')

            ax3.set_xlabel('Time (s)')
            ax3.set_ylabel('Global Sweep Number')
            ax3.set_title('Sniffing Bouts Across Experiments')
            ax3.grid(True, alpha=0.3)

            # Save figure to bytes
            buf3 = io.BytesIO()
            fig3.savefig(buf3, format='png', dpi=100, bbox_inches='tight')
            buf3.seek(0)
            plt.close(fig3)

            # Insert image into Excel
            img3 = XLImage(buf3)
            img3.width = 600
            img3.height = 360
            # Position to the right of apnea chart (column U)
            ws.add_image(img3, f"U{chart_start_row}")
            print(f"Added sniffing chart at U{chart_start_row}")


    def _add_sighs_chart(self, ws, header_row):
        """Add sigh timeline scatter plot to the sighs sheet."""
        import matplotlib.pyplot as plt
        from openpyxl.drawing.image import Image as XLImage
        import io
        import numpy as np

        print(f"_add_sighs_chart called for sheet with {ws.max_row} rows")

        # Find required columns
        exp_num_col = None
        global_sweep_col = None
        t_col = None

        for idx, cell in enumerate(header_row, start=1):
            if cell.value == 'experiment_number':
                exp_num_col = idx
            elif cell.value == 'global_sweep_number':
                global_sweep_col = idx
            elif cell.value == 't':
                t_col = idx

        if not all([exp_num_col, global_sweep_col, t_col]):
            print(f"Sighs sheet missing required columns for chart. Found columns: {[cell.value for cell in header_row]}")
            return

        # Read data from sheet
        sighs = []
        for row_idx in range(2, ws.max_row + 1):
            exp_num = ws.cell(row=row_idx, column=exp_num_col).value
            global_sweep = ws.cell(row=row_idx, column=global_sweep_col).value
            t = ws.cell(row=row_idx, column=t_col).value

            if all([exp_num is not None, global_sweep is not None, t is not None]):
                sighs.append({
                    'exp_num': int(exp_num),
                    'global_sweep': int(global_sweep),
                    't': float(t)
                })

        if not sighs:
            print("No sigh data found for chart generation")
            return

        # Create colormap for experiments
        n_experiments = max([s['exp_num'] for s in sighs])
        colors = plt.cm.tab10(np.linspace(0, 1, n_experiments))

        # Position chart at the top (starting at row 1)
        chart_start_row = 1

        # Create sigh scatter plot
        fig, ax = plt.subplots(figsize=(10, 6))

        # Plot each sigh as a yellow asterisk, colored by experiment
        for sigh in sighs:
            color = colors[sigh['exp_num'] - 1]
            ax.scatter(sigh['t'], sigh['global_sweep'],
                      marker='*', s=200, color='gold', edgecolors=color, linewidths=1.5)

        ax.set_xlabel('Time (s)')
        ax.set_ylabel('Global Sweep Number')
        ax.set_title('Sigh Events Across Experiments')
        ax.grid(True, alpha=0.3)

        # Save figure to bytes
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
        buf.seek(0)
        plt.close(fig)

        # Insert image into Excel
        img = XLImage(buf)
        img.width = 600
        img.height = 360
        ws.add_image(img, f"A{chart_start_row}")
        print(f"Added sigh chart at A{chart_start_row}")

