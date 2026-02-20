"""
Source Review Dialog — popup showing source documents with highlighted metadata cells.

Opens from right-click on a row in the Data Files table or the info button.
Shows metadata summary for the selected file/animal, with tabbed source document
previews and cell-level highlighting showing where each value was extracted from.

MVVM-compliant: no self.mw references, receives all data via constructor.
"""

import json
from pathlib import Path
from typing import Optional, Dict, List, Any

from PyQt6.QtCore import Qt, QMimeData, QSize
from PyQt6.QtGui import QColor, QFont, QKeySequence, QShortcut, QPixmap, QPainter, QIcon
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QTabWidget,
    QTableWidget, QTableWidgetItem, QHeaderView, QGroupBox,
    QGridLayout, QPushButton, QWidget, QScrollArea, QApplication,
)


# Field → highlight color mapping (consistent with old SourcePreviewWidget)
FIELD_COLORS = {
    "sex": QColor(100, 149, 237, 80),       # cornflower blue
    "strain": QColor(102, 205, 170, 80),     # medium aquamarine
    "power": QColor(255, 165, 0, 80),        # orange
    "animal_id": QColor(186, 85, 211, 80),   # medium orchid
    "stim_type": QColor(255, 99, 71, 80),    # tomato
    "group": QColor(144, 238, 144, 80),      # light green
    "experiment": QColor(135, 206, 250, 80), # light sky blue
}
DEFAULT_HIGHLIGHT = QColor(255, 255, 150, 80)  # pale yellow

# Shared tab styling for source document tabs (outer) and sheet tabs (inner)
_SOURCE_TAB_STYLE = """
    QTabWidget::pane {
        border: 1px solid #666;
        border-radius: 3px;
    }
    QTabBar::tab {
        padding: 5px 14px;
        margin-right: 2px;
        border: 1px solid #555;
        border-bottom: none;
        border-top-left-radius: 4px;
        border-top-right-radius: 4px;
        background: #2a2a2a;
        color: #999;
    }
    QTabBar::tab:selected {
        background: #3a3a4a;
        color: #ffffff;
        border-color: #7799cc;
        border-bottom: 2px solid #7799cc;
        font-weight: bold;
    }
    QTabBar::tab:hover:!selected {
        background: #333;
        color: #ccc;
    }
"""

_SHEET_TAB_STYLE = """
    QTabWidget::pane {
        border: 1px solid #555;
        border-radius: 2px;
    }
    QTabBar::tab {
        padding: 3px 10px;
        margin-right: 1px;
        border: 1px solid #444;
        border-top: none;
        border-bottom-left-radius: 3px;
        border-bottom-right-radius: 3px;
        background: #252525;
        color: #888;
        font-size: 11px;
    }
    QTabBar::tab:selected {
        background: #2e3548;
        color: #ffffff;
        border-color: #6688aa;
        border-top: 2px solid #6688aa;
        font-weight: bold;
    }
    QTabBar::tab:hover:!selected {
        background: #303030;
        color: #bbb;
    }
"""


class SourceReviewDialog(QDialog):
    """Popup dialog for reviewing source documents and metadata for a file/animal."""

    def __init__(
        self,
        file_name: str,
        animal_id: str,
        metadata: Dict[str, Any],
        sources: List[Dict[str, Any]],
        source_links: List[Dict[str, Any]],
        load_sheets_fn,
        parent: Optional[QWidget] = None,
    ):
        """
        Args:
            file_name: Display name of the recording file.
            animal_id: Animal ID for this row.
            metadata: Current metadata dict for this row (strain, sex, etc.).
            sources: List of source documents from ProjectService.get_sources().
            source_links: Source links for this animal from get_source_links(animal_id).
            load_sheets_fn: Callable(file_path) -> dict[sheet_name, list[list[str]]]
                            Loads sheet data from a source file.
            parent: Parent widget.
        """
        super().__init__(parent)
        self._file_name = file_name
        self._animal_id = animal_id
        self._metadata = metadata
        self._sources = sources
        self._source_links = source_links
        self._load_sheets = load_sheets_fn

        self.setWindowTitle(f"Source Review \u2014 {file_name} [Animal {animal_id}]")
        self.setMinimumSize(700, 500)
        self.resize(850, 600)

        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(8)

        # --- Metadata summary ---
        self._build_metadata_summary(layout)

        # --- Disagreement warnings ---
        self._build_disagreement_warnings(layout)

        # --- Source document tabs ---
        self._build_source_tabs(layout)

        # --- Legend + Close ---
        bottom = QHBoxLayout()
        self._legend = QLabel("")
        self._legend.setStyleSheet("font-size: 11px; color: #aaa;")
        self._build_legend()
        bottom.addWidget(self._legend, stretch=1)

        close_btn = QPushButton("Close")
        close_btn.setFixedWidth(80)
        close_btn.clicked.connect(self.close)
        bottom.addWidget(close_btn)
        layout.addLayout(bottom)

    # ------------------------------------------------------------------
    # Metadata summary bar
    # ------------------------------------------------------------------

    def _build_metadata_summary(self, parent_layout: QVBoxLayout):
        group = QGroupBox("Current Metadata")
        group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 1px solid #555;
                border-radius: 4px;
                margin-top: 8px;
                padding-top: 14px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 4px;
            }
        """)
        grid = QGridLayout(group)
        grid.setSpacing(4)
        grid.setHorizontalSpacing(4)

        fields = [
            ("Animal ID", "animal_id"), ("Strain", "strain"), ("Sex", "sex"),
            ("Stim Type", "stim_type"), ("Power", "power"), ("Experiment", "experiment"),
        ]

        for i, (label_text, key) in enumerate(fields):
            row, col = divmod(i, 3)
            lbl = QLabel(f"<b>{label_text}:</b>")
            lbl.setStyleSheet("color: #ccc;")
            val_text = str(self._metadata.get(key, "")) or "\u2014"

            # Check for disagreements on this field
            disagreement = self._get_disagreement(key)
            if disagreement:
                val_text += f"  \u26a0\ufe0f"

            val = QLabel(val_text)
            val.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)

            # Color-code value background to match the legend highlight colors
            field_color = FIELD_COLORS.get(key)
            if field_color and val_text != "\u2014":
                rgba = f"rgba({field_color.red()},{field_color.green()},{field_color.blue()},120)"
                val.setStyleSheet(
                    f"color: #fff; background: {rgba}; "
                    f"padding: 1px 6px; border-radius: 3px;"
                )
            else:
                val.setStyleSheet("color: #e0e0e0;")

            # Each field pair uses columns: col*3 (label), col*3+1 (value), col*3+2 (spacer)
            grid.addWidget(lbl, row, col * 3, Qt.AlignmentFlag.AlignRight)
            grid.addWidget(val, row, col * 3 + 1, Qt.AlignmentFlag.AlignLeft)

        # Set column stretches: label columns tight, value columns stretch,
        # small spacer between pairs
        for col in range(3):
            grid.setColumnStretch(col * 3, 0)      # label — no stretch
            grid.setColumnStretch(col * 3 + 1, 1)  # value — stretch
            if col < 2:
                grid.setColumnMinimumWidth(col * 3 + 2, 16)  # spacer between pairs

        parent_layout.addWidget(group)

    # ------------------------------------------------------------------
    # Disagreement warnings
    # ------------------------------------------------------------------

    @staticmethod
    def _normalize_value(value: str) -> str:
        """Normalize a value for comparison: strip units, whitespace, case."""
        import re
        v = str(value).strip().lower()
        # Strip common units (mW, mw, ms, hz, etc.)
        v = re.sub(r'\s*(mw|ms|hz|khz|mv|ua|ma)\s*$', '', v, flags=re.IGNORECASE)
        return v.strip()

    def _get_disagreement(self, field: str) -> Optional[List[Dict]]:
        """Check if source_links have genuinely conflicting values for a field.

        Deduplicates values that differ only by unit suffix (e.g. '2.5mW' vs '2.5')
        and compares across unique sources only (after dedup).
        """
        field_links = [l for l in self._source_links if l.get("field") == field]
        if len(field_links) < 2:
            return None

        # Deduplicate by (source_id, normalized_value) to avoid false positives
        # from the same source having both '2.5mW' and '2.5' in different columns
        seen = set()
        unique_links = []
        for link in field_links:
            key = (link.get("source_id"), self._normalize_value(link.get("value", "")))
            if key not in seen:
                seen.add(key)
                unique_links.append(link)

        if len(unique_links) < 2:
            return None

        # Compare normalized values across unique sources
        values = set(self._normalize_value(l.get("value", "")) for l in unique_links)
        if len(values) > 1:
            return unique_links
        return None

    def _build_disagreement_warnings(self, parent_layout: QVBoxLayout):
        """Add warning labels for fields with conflicting source values."""
        warnings = []
        for field in ("sex", "strain", "animal_id", "stim_type", "power", "experiment"):
            disagreement = self._get_disagreement(field)
            if disagreement:
                vals = []
                for link in disagreement:
                    src_id = link.get("source_id")
                    src_name = self._source_name(src_id)
                    vals.append(f'"{link.get("value", "")}" ({src_name})')
                warnings.append(f"\u26a0\ufe0f <b>{field}</b>: {' vs '.join(vals)}")

        if warnings:
            warning_label = QLabel("<br>".join(warnings))
            warning_label.setWordWrap(True)
            warning_label.setStyleSheet(
                "background-color: rgba(200, 150, 50, 40); "
                "border: 1px solid #c89632; border-radius: 4px; "
                "padding: 6px; color: #e0c060;"
            )
            parent_layout.addWidget(warning_label)

    def _source_name(self, source_id: int) -> str:
        """Get source display name by ID."""
        for s in self._sources:
            if s.get("source_id") == source_id or s.get("id") == source_id:
                name = s.get("name", "") or s.get("file_path", "")
                return Path(name).name if name else f"Source #{source_id}"
        return f"Source #{source_id}"

    # ------------------------------------------------------------------
    # Source document tabs
    # ------------------------------------------------------------------

    def _classify_sources(self, linked_sources):
        """Classify sources as primary (has file_name links for this file) or secondary.

        Primary sources directly reference the experiment file (e.g. pleth notes).
        Secondary sources have animal-level info only (e.g. surgery notes, mouse list).
        """
        primary = []
        secondary = []

        # Source IDs that have a file_name link matching this file
        file_stem = self._file_name
        for ext in ('.abf', '.smrx', '.edf', '.csv'):
            file_stem = file_stem.replace(ext, '')

        primary_src_ids = set()
        for link in self._source_links:
            if link.get("field") == "file_name":
                link_val = str(link.get("value", "")).strip()
                for ext in ('.abf', '.smrx', '.edf', '.csv'):
                    link_val = link_val.replace(ext, '')
                if link_val == file_stem:
                    primary_src_ids.add(link.get("source_id"))

        for src in linked_sources:
            src_id = src.get("source_id") or src.get("id")
            if src_id in primary_src_ids:
                primary.append(src)
            else:
                secondary.append(src)

        return primary, secondary

    def _deduplicate_sources(self, linked_sources: List[Dict]) -> List[Dict]:
        """Deduplicate sources that are the same file registered under different paths.

        The same notes file often gets registered multiple times when accessed via
        different path formats (UNC forward-slash, UNC backslash, mapped drive letter).
        Group by filename and keep only the version with the most source_links.
        """
        from collections import defaultdict

        # Group by filename (case-insensitive)
        by_name: dict = defaultdict(list)
        for src in linked_sources:
            src_path = src.get("file_path", "")
            name = Path(src_path).name.lower() if src_path else f"__{id(src)}"
            by_name[name].append(src)

        deduped = []
        merged_ids = {}  # source_id → canonical source_id
        for name, group in by_name.items():
            if len(group) == 1:
                deduped.append(group[0])
                continue

            # Pick the one with most links
            best = max(group, key=lambda s: sum(
                1 for l in self._source_links
                if l.get("source_id") == (s.get("source_id") or s.get("id"))
            ))
            deduped.append(best)
            best_id = best.get("source_id") or best.get("id")

            # Map duplicate source_ids to the canonical one so highlights still work
            for src in group:
                src_id = src.get("source_id") or src.get("id")
                if src_id != best_id:
                    merged_ids[src_id] = best_id

        # Remap source_links so all point to the canonical source
        if merged_ids:
            for link in self._source_links:
                old_id = link.get("source_id")
                if old_id in merged_ids:
                    link["source_id"] = merged_ids[old_id]

        return deduped

    def _build_source_tabs(self, parent_layout: QVBoxLayout):
        """Build tabbed view of source documents, primary first."""
        # Find which sources are referenced by links for this animal
        linked_source_ids = set(l.get("source_id") for l in self._source_links)
        linked_sources = [s for s in self._sources
                          if (s.get("source_id") or s.get("id")) in linked_source_ids]

        # Deduplicate sources registered under different path formats
        linked_sources = self._deduplicate_sources(linked_sources)

        if not linked_sources:
            placeholder = QLabel("No source documents linked to this animal.")
            placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
            placeholder.setStyleSheet("color: #888; padding: 20px;")
            parent_layout.addWidget(placeholder, stretch=1)
            return

        primary, secondary = self._classify_sources(linked_sources)

        # Header
        parts = []
        if primary:
            parts.append(f"{len(primary)} experiment notes")
        if secondary:
            parts.append(f"{len(secondary)} reference")
        source_label = QLabel(
            f"<b>Source Documents</b>  \u2014  {', '.join(parts)}"
        )
        source_label.setStyleSheet("color: #aaa; font-size: 11px; margin-top: 4px;")
        parent_layout.addWidget(source_label)

        source_tabs = QTabWidget()
        source_tabs.setStyleSheet(_SOURCE_TAB_STYLE)

        # Add primary sources first (experiment notes)
        for source in primary:
            self._add_source_tab(source_tabs, source, prefix="\u25c9 ")  # filled circle

        # Separator: visual gap via a disabled empty tab if both groups exist
        if primary and secondary:
            sep = QWidget()
            source_tabs.addTab(sep, "\u2502")  # vertical bar separator
            source_tabs.setTabEnabled(source_tabs.count() - 1, False)

        # Add secondary sources (surgery notes, mouse list, etc.)
        for source in secondary:
            self._add_source_tab(source_tabs, source, prefix="\u25cb ")  # open circle

        parent_layout.addWidget(source_tabs, stretch=1)

    @staticmethod
    def _make_dot_icon(colors: List[QColor], size: int = 16) -> QIcon:
        """Create a small icon with colored dots for a tab label."""
        dot_size = 6
        spacing = 2
        width = len(colors) * (dot_size + spacing) + spacing
        pm = QPixmap(max(width, size), size)
        pm.fill(QColor(0, 0, 0, 0))  # transparent
        painter = QPainter(pm)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        x = spacing
        y = (size - dot_size) // 2
        for color in colors:
            opaque = QColor(color.red(), color.green(), color.blue(), 220)
            painter.setBrush(opaque)
            painter.setPen(Qt.PenStyle.NoPen)
            painter.drawEllipse(x, y, dot_size, dot_size)
            x += dot_size + spacing
        painter.end()
        return QIcon(pm)

    def _add_source_tab(self, tab_widget: QTabWidget, source: Dict, prefix: str = ""):
        """Add a single source document as a tab."""
        src_id = source.get("source_id") or source.get("id")
        src_path = source.get("file_path", "")
        src_name = Path(src_path).name if src_path else f"Source #{src_id}"

        sheets = self._load_sheets(src_path)
        if not sheets:
            widget = QLabel(f"Could not load: {src_path}")
            widget.setAlignment(Qt.AlignmentFlag.AlignCenter)
            widget.setStyleSheet("color: #888; padding: 20px;")
        else:
            src_links = [l for l in self._source_links if l.get("source_id") == src_id]
            widget = self._build_sheet_tabs(sheets, src_links)

        tab_idx = tab_widget.addTab(widget, f"{prefix}{src_name}")

        # Build colored dot icon and tooltip from fields this source provides
        src_links = [l for l in self._source_links if l.get("source_id") == src_id]
        fields = sorted(set(l.get("field", "") for l in src_links if l.get("field")))
        if fields:
            tab_widget.setTabToolTip(tab_idx, f"Fields: {', '.join(fields)}")
            # Create colored dot icon matching the legend
            dot_colors = [FIELD_COLORS[f] for f in fields if f in FIELD_COLORS]
            if dot_colors:
                icon = self._make_dot_icon(dot_colors)
                tab_widget.setTabIcon(tab_idx, icon)
                tab_widget.setIconSize(QSize(len(dot_colors) * 8 + 4, 16))

    def _build_sheet_tabs(self, sheets: Dict[str, List[List[str]]],
                          links: List[Dict[str, Any]]) -> QWidget:
        """Build inner tab widget for sheets within a source document."""
        if len(sheets) == 1:
            # Single sheet — no need for inner tabs
            sheet_name, rows = next(iter(sheets.items()))
            return self._create_sheet_table(sheet_name, rows, links)

        inner_tabs = QTabWidget()
        inner_tabs.setTabPosition(QTabWidget.TabPosition.South)
        inner_tabs.setStyleSheet(_SHEET_TAB_STYLE)

        # Determine which sheets have linked data
        linked_sheets = set()
        for link in links:
            loc = self._parse_location(link.get("location", ""))
            sheet = loc.get("sheet")
            if sheet:
                linked_sheets.add(sheet)

        focus_tab_idx = 0
        focus_found = False

        for idx, (sheet_name, rows) in enumerate(sheets.items()):
            table = self._create_sheet_table(sheet_name, rows, links)
            is_linked = sheet_name in linked_sheets

            # Mark linked sheet with a star prefix
            tab_label = f"\u2605 {sheet_name}" if is_linked else sheet_name
            inner_tabs.addTab(table, tab_label)

            if is_linked:
                # Permanent color on the linked tab so it's always visible
                inner_tabs.tabBar().setTabTextColor(
                    idx, QColor(120, 180, 255)  # bright blue
                )
                inner_tabs.setTabToolTip(idx, f"{sheet_name} — has linked metadata")
                if not focus_found:
                    focus_tab_idx = idx
                    focus_found = True

        if focus_found:
            inner_tabs.setCurrentIndex(focus_tab_idx)

        return inner_tabs

    def _create_sheet_table(self, sheet_name: str, rows: List[List[str]],
                            links: List[Dict[str, Any]]) -> QTableWidget:
        """Create a QTableWidget for one sheet with highlighted cells."""
        if not rows:
            table = QTableWidget(0, 0)
            return table

        n_rows = len(rows)
        n_cols = max(len(r) for r in rows) if rows else 0

        table = QTableWidget(n_rows, n_cols)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        # Allow multi-selection for copy-paste
        table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)

        header = table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        header.setStretchLastSection(True)
        table.verticalHeader().setDefaultSectionSize(22)

        # Populate cells
        for r, row_data in enumerate(rows):
            for c, val in enumerate(row_data):
                item = QTableWidgetItem(str(val) if val is not None else "")
                # Allow selection but not editing
                item.setFlags(
                    Qt.ItemFlag.ItemIsEnabled
                    | Qt.ItemFlag.ItemIsSelectable
                )
                table.setItem(r, c, item)

        # Apply highlights from source links
        scroll_to_item = None
        for link in links:
            loc = self._parse_location(link.get("location", ""))
            link_sheet = loc.get("sheet", "")

            # Match this link to the current sheet
            if link_sheet and link_sheet != sheet_name:
                continue

            row_idx = loc.get("row")
            col_idx = loc.get("col")
            if row_idx is None or col_idx is None:
                continue
            if row_idx >= n_rows or col_idx >= n_cols:
                continue

            item = table.item(row_idx, col_idx)
            if not item:
                continue

            field = link.get("field", "")
            color = FIELD_COLORS.get(field, DEFAULT_HIGHLIGHT)
            item.setBackground(color)

            animal_id = link.get("animal_id", "")
            value = link.get("value", "")
            confidence = link.get("confidence", 0)
            tooltip = f"{field}: {value}"
            if animal_id:
                tooltip = f"Animal {animal_id} | {tooltip}"
            if confidence:
                tooltip += f" ({confidence:.0%})"
            item.setToolTip(tooltip)

            # Track first highlighted item for auto-scroll
            if scroll_to_item is None and animal_id == self._animal_id:
                scroll_to_item = item

        # Auto-scroll to animal's row
        if scroll_to_item:
            table.scrollToItem(scroll_to_item, QTableWidget.ScrollHint.PositionAtCenter)

        # Enable Ctrl+C copy
        _install_copy_shortcut(table)

        return table

    # ------------------------------------------------------------------
    # Legend
    # ------------------------------------------------------------------

    def _build_legend(self):
        """Build color legend from fields present in source links."""
        fields = sorted(set(l.get("field", "") for l in self._source_links if l.get("field")))
        if not fields:
            return

        parts = []
        for field in fields[:8]:
            color = FIELD_COLORS.get(field, DEFAULT_HIGHLIGHT)
            # Use full opacity for legend swatches so they're readable
            rgb = f"rgba({color.red()},{color.green()},{color.blue()},180)"
            parts.append(
                f'<span style="background:{rgb}; color:#fff; padding:2px 8px; '
                f'border-radius:3px; margin:0 2px; font-weight:bold;">{field}</span>'
            )

        self._legend.setText(
            '<span style="color:#ccc; font-size:12px;">Highlights: </span>' + " ".join(parts)
        )
        self._legend.setStyleSheet("font-size: 12px; color: #ccc; padding: 2px 0;")

    # ------------------------------------------------------------------
    # Utilities
    # ------------------------------------------------------------------

    @staticmethod
    def _parse_location(location) -> Dict[str, Any]:
        """Parse location JSON string or dict."""
        if not location:
            return {}
        if isinstance(location, dict):
            return location
        try:
            return json.loads(location)
        except (json.JSONDecodeError, TypeError):
            return {}


def _install_copy_shortcut(table: QTableWidget):
    """Install Ctrl+C shortcut to copy selected cells as tab-separated text."""
    def _copy():
        selection = table.selectedRanges()
        if not selection:
            return
        lines = []
        for sel_range in selection:
            for row in range(sel_range.topRow(), sel_range.bottomRow() + 1):
                cols = []
                for col in range(sel_range.leftColumn(), sel_range.rightColumn() + 1):
                    item = table.item(row, col)
                    cols.append(item.text() if item else "")
                lines.append("\t".join(cols))
        text = "\n".join(lines)
        QApplication.clipboard().setText(text)

    shortcut = QShortcut(QKeySequence.StandardKey.Copy, table)
    shortcut.activated.connect(_copy)


def load_source_sheets(file_path: str) -> Dict[str, List[List[str]]]:
    """Load sheet data from a source file (Excel/CSV) for preview.

    Standalone function — can be passed as load_sheets_fn to the dialog.
    """
    path = Path(file_path)
    if not path.exists():
        return {}

    try:
        if path.suffix.lower() in ('.xlsx', '.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            sheets = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append([str(c) if c is not None else "" for c in row])
                sheets[sheet_name] = rows
            wb.close()
            return sheets
        elif path.suffix.lower() == '.csv':
            import csv
            with open(str(path), 'r', encoding='utf-8', errors='replace') as f:
                reader = csv.reader(f)
                rows = [list(row) for row in reader]
            return {path.stem: rows}
    except Exception as e:
        print(f"[source-review] Error loading {file_path}: {e}")
        return {}

    return {}
